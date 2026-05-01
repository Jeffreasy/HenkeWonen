from __future__ import annotations

import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "DATA"
OUT_DIR = ROOT / "docs"
CONVERTED_DIR = ROOT / ".audit-cache" / "converted-xls"


PRICE_TYPE_KEYWORDS: list[tuple[str, str]] = [
    ("netto", "net_purchase"),
    ("net purchase", "net_purchase"),
    ("commissie", "commission"),
    ("commisie", "commission"),
    ("commission", "commission"),
    ("inkoop", "purchase"),
    ("purchase", "purchase"),
    ("advies", "advice_retail"),
    ("consumer", "advice_retail"),
    ("verkoop", "retail"),
    ("retail", "retail"),
    ("pallet", "pallet"),
    ("trailer", "trailer"),
    ("rolprijs", "roll"),
    ("roll", "roll"),
    ("coupage", "cut_length"),
    ("verpakking", "package"),
    ("package", "package"),
    ("pak", "package"),
    ("trede", "step"),
    ("step", "step"),
]

UNIT_KEYWORDS: list[tuple[str, str]] = [
    ("m²", "m2"),
    ("m2", "m2"),
    ("m¹", "m1"),
    ("m1", "m1"),
    ("meter", "meter"),
    ("lengte", "meter"),
    ("stuk", "piece"),
    ("piece", "piece"),
    ("verpakking", "package"),
    ("package", "package"),
    ("pak", "pack"),
    ("roll", "roll"),
    ("rol", "roll"),
    ("pallet", "pallet"),
    ("trailer", "trailer"),
    ("trede", "step"),
    ("step", "step"),
    ("liter", "liter"),
    ("kg", "kg"),
]

HEADER_KEYWORDS = [
    "art",
    "artikel",
    "ean",
    "sku",
    "code",
    "omschrijving",
    "description",
    "product",
    "kleur",
    "color",
    "kwaliteit",
    "quality",
    "design",
    "prijs",
    "price",
    "breedte",
    "width",
    "collectie",
    "collection",
]

CODE_HEADER_RE = re.compile(
    r"\b(art\.?\s*nr\.?|artikelnummer|artikelnr\.?|ean(?:\s*code)?|sku|supplier\s*code|leverancier.*code|commercial\s*code|sap\s*codes?)\b",
    re.IGNORECASE,
)

PRICE_HEADER_RE = re.compile(
    r"(prijs|price|verkoop|inkoop|consumer|retail|purchase|commi?s?s?ie|trailerprijs|palletprijs|rolprijs|coupage)",
    re.IGNORECASE,
)

QUANTITY_HEADER_RE = re.compile(
    r"(aantal\s+(?:pak|pakken|pakker|panelen|m2|m²|per)|pakken|pakker|panelen\s+per\s+pak|planken\s+per\s+pak|pakinhoud|plinten\s+per|afname|besteleenheid|bundel|packs|panels|m2\s+per\s+pak|m²\s+per\s+pak|per\s+pallet|per\s+trailer)",
    re.IGNORECASE,
)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_header(value: Any) -> str:
    return clean_text(value).replace("\n", " ").strip()


def safe_sample(value: Any) -> str:
    text = clean_text(value)
    if len(text) > 80:
        text = text[:77] + "..."
    return text


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def infer_price_type(header: str) -> str:
    lower = header.lower()
    for keyword, price_type in PRICE_TYPE_KEYWORDS:
        if keyword in lower:
            return price_type
    return "manual"


def infer_unit(header: str) -> str:
    lower = header.lower()
    for keyword, unit in UNIT_KEYWORDS:
        if keyword in lower:
            return unit
    return "custom"


def infer_vat_mode(header: str) -> str:
    lower = header.lower()
    if "incl" in lower and ("btw" in lower or "vat" in lower):
        return "inclusive"
    if "excl" in lower and ("btw" in lower or "vat" in lower):
        return "exclusive"
    return "unknown"


def is_price_header(header: str) -> bool:
    return bool(PRICE_HEADER_RE.search(header))


def is_quantity_header(header: str) -> bool:
    lower = header.lower()
    if "aantal kleuren" in lower:
        return False
    return bool(QUANTITY_HEADER_RE.search(header)) and not is_price_header(header)


def infer_quantity_kind(header: str) -> str:
    lower = header.lower()
    if "trailer" in lower:
        return "trailerQuantity"
    if "pallet" in lower:
        return "palletQuantity"
    if "bundel" in lower:
        return "bundleSize"
    if "besteleenheid" in lower or "afname" in lower:
        return "orderUnit"
    if "pak" in lower or "pack" in lower:
        return "packageQuantity"
    return "quantity"


def is_numeric_like(value: Any) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    text = clean_text(value)
    if not text:
        return False
    normalized = text.replace("€", "").replace(" ", "").replace(".", "").replace(",", ".")
    try:
        float(normalized)
        return True
    except ValueError:
        return False


def row_values(ws, row_idx: int, max_col: int) -> list[Any]:
    return [ws.cell(row_idx, col_idx).value for col_idx in range(1, max_col + 1)]


def sheet_rows(ws, min_row: int, max_row: int | None, max_col: int):
    yield from ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    )


def nonempty(values: list[Any]) -> list[Any]:
    return [value for value in values if clean_text(value) != ""]


def header_score(values: list[Any]) -> float:
    texts = [normalize_header(value).lower() for value in values if normalize_header(value)]
    if not texts:
        return 0
    joined = " | ".join(texts)
    keyword_hits = sum(1 for keyword in HEADER_KEYWORDS if keyword in joined)
    price_hits = sum(1 for text in texts if PRICE_HEADER_RE.search(text))
    code_hits = sum(1 for text in texts if CODE_HEADER_RE.search(text))
    return len(texts) + keyword_hits * 2 + price_hits * 2 + code_hits * 2


def detect_header_row(ws, max_scan_rows: int, max_col: int) -> int | None:
    candidates: list[tuple[float, int]] = []
    scan_to = min(ws.max_row or 0, max_scan_rows)
    for offset, row in enumerate(sheet_rows(ws, 1, scan_to, max_col), start=1):
        row_idx = offset
        values = list(row)
        score = header_score(values)
        if score:
            candidates.append((score, row_idx))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    score, row_idx = candidates[0]
    return row_idx if score >= 6 else None


def row_kind(values: list[Any], code_indexes: list[int], price_indexes: list[int]) -> str:
    filled = nonempty(values)
    if not filled:
        return "empty"
    has_code = any(clean_text(values[index]) for index in code_indexes if index < len(values))
    has_price = any(is_numeric_like(values[index]) for index in price_indexes if index < len(values))
    numeric_count = sum(1 for value in filled if is_numeric_like(value))
    if len(filled) <= 2 and not has_code and not has_price and numeric_count == 0:
        return "section"
    if has_code or has_price or len(filled) >= 3:
        return "product"
    return "warning"


def summarize_sheet(path: Path, ws) -> dict[str, Any]:
    max_col = min(ws.max_column or 1, 80)
    header_row = detect_header_row(ws, 100, max_col)
    headers: list[str] = []
    if header_row:
        header_values = next(sheet_rows(ws, header_row, header_row, max_col), [])
        headers = [normalize_header(value) for value in header_values]
        while headers and headers[-1] == "":
            headers.pop()

    code_indexes = [idx for idx, header in enumerate(headers) if CODE_HEADER_RE.search(header)]
    price_indexes = [idx for idx, header in enumerate(headers) if is_price_header(header)]
    quantity_indexes = [idx for idx, header in enumerate(headers) if is_quantity_header(header)]
    price_columns = [
        {
            "index": idx + 1,
            "header": headers[idx],
            "priceType": infer_price_type(headers[idx]),
            "priceUnit": infer_unit(headers[idx]),
            "vatMode": infer_vat_mode(headers[idx]),
        }
        for idx in price_indexes
    ]
    quantity_columns = [
        {
            "index": idx + 1,
            "header": headers[idx],
            "kind": infer_quantity_kind(headers[idx]),
        }
        for idx in quantity_indexes
    ]

    row_counts = Counter()
    section_samples: list[str] = []
    code_samples: dict[str, list[str]] = defaultdict(list)
    numeric_code_cells: list[dict[str, Any]] = []
    leading_dot_codes: list[dict[str, Any]] = []
    first_product_rows: list[dict[str, Any]] = []

    start = (header_row + 1) if header_row else 1
    max_iter_row = min(ws.max_row or 0, 50000)
    empty_streak = 0
    last_observed_row = 0
    for row_idx, row in enumerate(sheet_rows(ws, start, max_iter_row, max_col), start=start):
        values = list(row)
        kind = row_kind(values, code_indexes, price_indexes)
        if kind == "empty":
            empty_streak += 1
            if row_idx > start + 100 and empty_streak >= 500:
                break
            continue
        empty_streak = 0
        last_observed_row = row_idx
        row_counts[kind] += 1
        filled = nonempty(values)
        if kind == "section" and len(section_samples) < 8:
            section_samples.append(safe_sample(filled[0]))
        if kind == "product" and len(first_product_rows) < 3:
            product_summary = {}
            for idx, header in enumerate(headers[: len(values)]):
                if header and (
                    idx in code_indexes
                    or idx in price_indexes
                    or idx in quantity_indexes
                    or len(product_summary) < 8
                ):
                    product_summary[header] = safe_sample(values[idx])
            first_product_rows.append(product_summary)
        for idx in code_indexes:
            if idx >= len(values):
                continue
            value = values[idx]
            text = clean_text(value)
            if not text:
                continue
            header = headers[idx]
            if len(code_samples[header]) < 6:
                code_samples[header].append(text)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                numeric_code_cells.append({"row": row_idx, "header": header, "value": text})
            if text.startswith("."):
                leading_dot_codes.append({"row": row_idx, "header": header, "value": text})

    return {
        "sheetName": ws.title,
        "maxRow": ws.max_row,
        "maxColumn": ws.max_column,
        "lastObservedRow": last_observed_row,
        "headerRow": header_row,
        "headers": headers,
        "codeColumns": [headers[idx] for idx in code_indexes],
        "priceColumns": price_columns,
        "quantityColumns": quantity_columns,
        "rowCounts": dict(row_counts),
        "sectionSamples": section_samples,
        "codeSamples": dict(code_samples),
        "numericCodeCells": numeric_code_cells[:20],
        "leadingDotCodes": leading_dot_codes[:20],
        "firstProductRows": first_product_rows,
    }


def summarize_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = [summarize_sheet(path, wb[sheet_name]) for sheet_name in wb.sheetnames]
    finally:
        wb.close()
    return {
        "path": str(path.relative_to(ROOT)),
        "fileName": path.name,
        "extension": path.suffix.lower(),
        "sizeBytes": path.stat().st_size,
        "sha256": sha256(path),
        "sheets": sheets,
    }


def converted_xlsx_for(path: Path) -> Path | None:
    if path.suffix.lower() != ".xls":
        return path
    candidate = CONVERTED_DIR / f"{path.stem}.xlsx"
    return candidate if candidate.exists() else None


def build_markdown(audit: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append("# Henke Wonen Data Audit")
    lines.append("")
    lines.append("Deze audit is gegenereerd uit de lokale `DATA` map. Persoonsdata uit klantbestanden is niet als detail overgenomen.")
    lines.append("")
    lines.append("## Overzicht")
    lines.append("")
    lines.append(f"- Bronbestanden gevonden: {audit['fileCount']}")
    lines.append(f"- Geanalyseerde werkboeken: {audit['analyzedWorkbookCount']}")
    lines.append(f"- Niet geanalyseerd: {len(audit['skippedFiles'])}")
    lines.append(f"- Totaal sheets: {audit['sheetCount']}")
    lines.append(f"- Geschatte productrijen: {audit['estimatedProductRows']}")
    lines.append("")

    if audit["skippedFiles"]:
        lines.append("## Overgeslagen bestanden")
        lines.append("")
        for item in audit["skippedFiles"]:
            lines.append(f"- `{item['path']}`: {item['reason']}")
        lines.append("")

    lines.append("## Werkboeken en sheets")
    lines.append("")
    lines.append("| Bestand | Sheet | Header row | Productrijen | Sectierijen | Prijskolommen | Logistieke kolommen | Codekolommen |")
    lines.append("| --- | --- | ---: | ---: | ---: | ---: | ---: | --- |")
    for workbook in audit["workbooks"]:
        for sheet in workbook["sheets"]:
            rows = sheet["rowCounts"]
            code_cols = ", ".join(sheet["codeColumns"][:4])
            if len(sheet["codeColumns"]) > 4:
                code_cols += ", ..."
            lines.append(
                f"| `{workbook['fileName']}` | `{sheet['sheetName']}` | "
                f"{sheet['headerRow'] or ''} | {rows.get('product', 0)} | "
                f"{rows.get('section', 0)} | {len(sheet['priceColumns'])} | "
                f"{len(sheet['quantityColumns'])} | {code_cols} |"
            )
    lines.append("")

    lines.append("## Prijskolommen")
    lines.append("")
    lines.append("| Header | Aantal | Inferred priceType | Inferred priceUnit | Inferred vatMode |")
    lines.append("| --- | ---: | --- | --- | --- |")
    for header, count in audit["priceHeaderCounts"].most_common():
        meta = audit["priceHeaderMeta"][header]
        lines.append(
            f"| `{header}` | {count} | `{meta['priceType']}` | `{meta['priceUnit']}` | `{meta['vatMode']}` |"
        )
    lines.append("")

    lines.append("## Logistieke en verpakkingskolommen")
    lines.append("")
    lines.append("| Header | Aantal | Inferred veld |")
    lines.append("| --- | ---: | --- |")
    for header, count in audit["quantityHeaderCounts"].most_common():
        kind = audit["quantityHeaderMeta"][header]["kind"]
        lines.append(f"| `{header}` | {count} | `{kind}` |")
    lines.append("")

    lines.append("## Import-risico's")
    lines.append("")
    for risk in audit["risks"]:
        lines.append(f"- {risk}")
    lines.append("")

    lines.append("## Duplicaten op bestandsnaam")
    lines.append("")
    if audit["duplicateFileNames"]:
        for name, paths in audit["duplicateFileNames"].items():
            lines.append(f"- `{name}`")
            for item in paths:
                lines.append(f"  - `{item}`")
    else:
        lines.append("- Geen dubbele bestandsnamen gevonden.")
    lines.append("")

    return "\n".join(lines)


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    files = sorted(
        [
            path
            for path in DATA_DIR.rglob("*")
            if path.is_file() and path.suffix.lower() in {".xlsx", ".xls"}
        ]
    )

    workbooks = []
    skipped = []
    for source_path in files:
        analysis_path = converted_xlsx_for(source_path)
        if analysis_path is None:
            skipped.append(
                {
                    "path": str(source_path.relative_to(ROOT)),
                    "reason": "Legacy .xls has not been converted to .xlsx in .audit-cache/converted-xls.",
                }
            )
            continue
        try:
            summary = summarize_workbook(analysis_path)
            summary["sourcePath"] = str(source_path.relative_to(ROOT))
            summary["analysisPath"] = str(analysis_path.relative_to(ROOT))
            workbooks.append(summary)
        except Exception as exc:  # noqa: BLE001 - audit should continue across bad workbooks
            skipped.append(
                {
                    "path": str(source_path.relative_to(ROOT)),
                    "reason": f"{type(exc).__name__}: {exc}",
                }
            )

    price_header_counts: Counter[str] = Counter()
    price_header_meta: dict[str, dict[str, str]] = {}
    quantity_header_counts: Counter[str] = Counter()
    quantity_header_meta: dict[str, dict[str, str]] = {}
    sheet_count = 0
    estimated_product_rows = 0
    numeric_code_cells = []
    leading_dot_codes = []
    section_sheet_count = 0
    unknown_vat_columns = []

    for workbook in workbooks:
        sheet_count += len(workbook["sheets"])
        for sheet in workbook["sheets"]:
            estimated_product_rows += sheet["rowCounts"].get("product", 0)
            if sheet["rowCounts"].get("section", 0):
                section_sheet_count += 1
            numeric_code_cells.extend(
                {
                    **cell,
                    "fileName": workbook["fileName"],
                    "sheetName": sheet["sheetName"],
                }
                for cell in sheet["numericCodeCells"]
            )
            leading_dot_codes.extend(
                {
                    **cell,
                    "fileName": workbook["fileName"],
                    "sheetName": sheet["sheetName"],
                }
                for cell in sheet["leadingDotCodes"]
            )
            for column in sheet["priceColumns"]:
                header = column["header"]
                price_header_counts[header] += 1
                price_header_meta[header] = {
                    "priceType": column["priceType"],
                    "priceUnit": column["priceUnit"],
                    "vatMode": column["vatMode"],
                }
                if column["vatMode"] == "unknown":
                    unknown_vat_columns.append(
                        {
                            "fileName": workbook["fileName"],
                            "sheetName": sheet["sheetName"],
                            "header": header,
                        }
                    )
            for column in sheet["quantityColumns"]:
                header = column["header"]
                quantity_header_counts[header] += 1
                quantity_header_meta[header] = {
                    "kind": column["kind"],
                }

    by_name: dict[str, list[str]] = defaultdict(list)
    for path in files:
        by_name[path.name].append(str(path.relative_to(ROOT)))
    duplicate_file_names = {
        name: paths for name, paths in by_name.items() if len(paths) > 1
    }

    risks = []
    if any(path.suffix.lower() == ".xls" for path in files):
        risks.append(".xls support is verplicht; Interfloor is legacy Excel en moet via parser/conversiepad blijven werken.")
    if numeric_code_cells:
        risks.append(
            f"{len(numeric_code_cells)} codecellen kwamen als numeriek uit Excel; import moet codes expliciet naar string serialiseren."
        )
    if leading_dot_codes:
        risks.append(
            f"{len(leading_dot_codes)} codevoorbeelden beginnen met een punt; trimming/number-casting zou deze beschadigen."
        )
    if unknown_vat_columns:
        risks.append(
            f"{len(unknown_vat_columns)} prijskolom-observaties hebben onbekende btw-modus; import-preview moet mapping bevestiging vragen."
        )
    if section_sheet_count:
        risks.append(
            f"{section_sheet_count} sheets bevatten vermoedelijke sectierijen; sectionLabel/subcategory moet doorlopen naar volgende productregels."
        )
    if duplicate_file_names:
        risks.append(
            f"{len(duplicate_file_names)} bestandsnamen komen op meerdere locaties voor; deduplicatie mag niet op bestandsnaam gebeuren."
        )
    if quantity_header_counts:
        risks.append(
            f"{sum(quantity_header_counts.values())} logistieke/verpakkingskolom-observaties gevonden; import moet deze scheiden van productPrices."
        )

    audit = {
        "fileCount": len(files),
        "analyzedWorkbookCount": len(workbooks),
        "skippedFiles": skipped,
        "sheetCount": sheet_count,
        "estimatedProductRows": estimated_product_rows,
        "workbooks": workbooks,
        "priceHeaderCounts": price_header_counts,
        "priceHeaderMeta": price_header_meta,
        "quantityHeaderCounts": quantity_header_counts,
        "quantityHeaderMeta": quantity_header_meta,
        "numericCodeCells": numeric_code_cells[:100],
        "leadingDotCodes": leading_dot_codes[:100],
        "unknownVatColumns": unknown_vat_columns[:200],
        "duplicateFileNames": duplicate_file_names,
        "risks": risks,
    }

    serializable = {
        **audit,
        "priceHeaderCounts": dict(price_header_counts),
        "quantityHeaderCounts": dict(quantity_header_counts),
    }
    (OUT_DIR / "data-audit.json").write_text(
        json.dumps(serializable, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    (OUT_DIR / "data-audit.md").write_text(build_markdown(audit), encoding="utf-8")
    print(
        json.dumps(
            {
                "fileCount": audit["fileCount"],
                "analyzedWorkbookCount": audit["analyzedWorkbookCount"],
                "skipped": len(skipped),
                "sheetCount": sheet_count,
                "estimatedProductRows": estimated_product_rows,
                "risks": risks,
            },
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
