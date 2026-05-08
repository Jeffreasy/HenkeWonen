from __future__ import annotations

import fnmatch
import argparse
import json
import math
import os
import re
import shutil
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import sys

sys.path.append(str(Path(__file__).resolve().parent))

from audit_excel_data import DATA_DIR, ROOT, clean_text, safe_sample  # noqa: E402
from build_catalog_import import converted_xlsx_for, parse_workbook  # noqa: E402


AUDIT_DIR = ROOT / "docs" / "audit"
INVENTORY_OUT = AUDIT_DIR / "02_source_inventory.md"
RECONCILIATION_OUT = AUDIT_DIR / "03_reconciliation_report.xlsx"
TENANT_SLUG = "henke-wonen"
DEVIATION_THRESHOLD = 0.005
PRODUCTION_CONVEX_DEPLOYMENT = "prod:accomplished-kangaroo-354"
PRODUCTION_CONVEX_NAME = "accomplished-kangaroo-354"
PRODUCTION_CONVEX_URL = "https://accomplished-kangaroo-354.eu-west-1.convex.cloud"


@dataclass
class ToolContext:
    tenant_slug: str
    target: str
    target_option: str | None
    convex_deployment: str | None
    convex_url: str | None
    env_file: Path | None
    env_file_loaded: bool
    env_file_explicit: bool
    skip_env_file: bool
    source_filters: list[str]
    no_write: bool


@dataclass
class SourceAudit:
    path: Path
    file_name: str
    profile_name: str | None
    excel_articles: set[str]
    matched_articles: set[str]
    coverage_pct: float | None
    status: str
    deviations: list[dict[str, Any]]
    untraced: list[dict[str, Any]]
    vat_open: int
    verdict: str


def clean_env_value(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = value.strip()
    if len(stripped) >= 2 and stripped[0] == stripped[-1] and stripped[0] in {"'", '"'}:
        return stripped[1:-1]
    return stripped


def load_env_file(path: Path) -> bool:
    if not path.exists():
        return False

    for line in path.read_text(encoding="utf-8").splitlines():
        trimmed = line.strip()
        if not trimmed or trimmed.startswith("#") or "=" not in trimmed:
            continue
        key, value = trimmed.split("=", 1)
        if key and key not in os.environ:
            os.environ[key] = clean_env_value(value) or ""

    return True


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Reconcile local catalog source workbooks against Convex catalog data."
    )
    parser.add_argument("--env-file")
    parser.add_argument("--no-env-file", action="store_true")
    parser.add_argument("--production", action="store_true")
    parser.add_argument("--target")
    parser.add_argument("--tenant", default=TENANT_SLUG)
    parser.add_argument("--source", action="append", default=[])
    parser.add_argument("--no-write", action="store_true")
    return parser.parse_args(argv)


def infer_target(target_option: str | None, convex_deployment: str | None, convex_url: str | None) -> str:
    normalized_target = (target_option or "").lower()
    if normalized_target in {"production", "prod"}:
        return "production"
    if normalized_target in {"development", "dev", "local"}:
        return "development"
    if convex_deployment == PRODUCTION_CONVEX_DEPLOYMENT or convex_url == PRODUCTION_CONVEX_URL:
        return "production"
    if convex_deployment and (convex_deployment.startswith("dev:") or convex_deployment == "local"):
        return "development"
    return "unknown"


def has_literal_newline_artifact(value: str | None) -> bool:
    return bool(value and ("\\r" in value or "\\n" in value or "\r" in value or "\n" in value))


def build_context(argv: list[str]) -> ToolContext:
    args = parse_args(argv)
    env_file_option = args.env_file or os.environ.get("CATALOG_ENV_FILE")
    env_file_explicit = bool(env_file_option)
    env_path = None if args.no_env_file else ROOT / (env_file_option or ".env.local")
    env_file_loaded = load_env_file(env_path) if env_path else False
    convex_deployment = clean_env_value(os.environ.get("CONVEX_DEPLOYMENT"))
    convex_url = clean_env_value(os.environ.get("PUBLIC_CONVEX_URL"))
    target = infer_target(args.target, convex_deployment, convex_url)
    production_flag = args.production or (args.target or "").lower() in {"production", "prod"}

    if not convex_url:
        raise SystemExit("PUBLIC_CONVEX_URL ontbreekt. Gebruik --env-file of --no-env-file met shell-env.")

    for key, value in {
        "CONVEX_DEPLOYMENT": convex_deployment,
        "PUBLIC_CONVEX_URL": convex_url,
        "PUBLIC_CONVEX_HTTP_ACTIONS_URL": clean_env_value(os.environ.get("PUBLIC_CONVEX_HTTP_ACTIONS_URL")),
        "CONVEX_SITE_URL": clean_env_value(os.environ.get("CONVEX_SITE_URL")),
    }.items():
        if has_literal_newline_artifact(value):
            raise SystemExit(f"{key} bevat een literal newline artifact. Trek/schrijf de env opnieuw schoon.")

    if target == "production":
        if not env_file_explicit and not args.no_env_file:
            raise SystemExit(
                "Reconciliation op production vereist --env-file/ CATALOG_ENV_FILE of bewust --no-env-file met shell-env."
            )
        if not production_flag:
            raise SystemExit("Reconciliation wijst naar production, maar mist --production of --target=production.")
        if convex_deployment != PRODUCTION_CONVEX_DEPLOYMENT or convex_url != PRODUCTION_CONVEX_URL:
            raise SystemExit(
                f"Reconciliation mag alleen naar {PRODUCTION_CONVEX_DEPLOYMENT} / {PRODUCTION_CONVEX_URL}."
            )
    elif production_flag:
        raise SystemExit(f"Reconciliation kreeg een production-flag, maar de geladen env is target={target}.")

    if target == "unknown":
        raise SystemExit("Reconciliation target is onbekend. Zet CONVEX_DEPLOYMENT of --target expliciet.")

    return ToolContext(
        tenant_slug=args.tenant,
        target=target,
        target_option=args.target,
        convex_deployment=convex_deployment,
        convex_url=convex_url,
        env_file=env_path,
        env_file_loaded=env_file_loaded,
        env_file_explicit=env_file_explicit,
        skip_env_file=args.no_env_file,
        source_filters=[item.lower() for item in args.source],
        no_write=args.no_write,
    )


def convex_deployment_arg(context: ToolContext) -> str | None:
    deployment = context.convex_deployment
    if context.target == "production":
        return PRODUCTION_CONVEX_NAME
    if not deployment:
        return None
    if deployment in {"dev", "prod", "local"}:
        return deployment
    return None


def matches_source_filter(path: Path, filters: list[str]) -> bool:
    if not filters:
        return True
    relative = str(path.relative_to(ROOT)).replace("\\", "/").lower()
    name = path.name.lower()
    return any(filter_value in name or filter_value in relative for filter_value in filters)


def source_files(context: ToolContext) -> list[Path]:
    return sorted(
        path
        for path in DATA_DIR.rglob("*")
        if path.is_file()
        and path.suffix.lower() in {".xlsx", ".xls"}
        and matches_source_filter(path, context.source_filters)
    )


def parse_json_output(output: str) -> Any:
    try:
        return json.loads(output)
    except json.JSONDecodeError:
        start = output.find("{")
        end = output.rfind("}")
        if start == -1 or end == -1 or end < start:
            raise
        return json.loads(output[start : end + 1])


def convex_run(context: ToolContext, query: str) -> Any:
    npx = shutil.which("npx")
    if not npx:
        raise SystemExit("npx was not found; cannot read Convex snapshot.")
    inline_query = " ".join(line.strip() for line in query.splitlines() if line.strip())
    command = [npx, "convex", "run"]
    deployment_arg = convex_deployment_arg(context)
    if deployment_arg:
        command.extend(["--deployment", deployment_arg])
    command.extend(["--inline-query", inline_query])
    env = os.environ.copy()
    if context.convex_deployment:
        env["CONVEX_DEPLOYMENT"] = context.convex_deployment
    if context.convex_url:
        env["PUBLIC_CONVEX_URL"] = context.convex_url
    result = subprocess.run(
        command,
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
        env=env,
    )
    if result.returncode != 0:
        raise SystemExit(result.stderr.strip() or result.stdout.strip())
    return parse_json_output(result.stdout)


def convex_page(context: ToolContext, table_name: str, mapper: str, cursor: str | None) -> dict[str, Any]:
    cursor_literal = json.dumps(cursor)
    tenant_slug = json.dumps(context.tenant_slug)
    query = f"""
const tenant = await ctx.db.query('tenants').withIndex('by_slug', q => q.eq('slug', {tenant_slug})).unique();
const result = await ctx.db
  .query('{table_name}')
  .withIndex('by_tenant', q => q.eq('tenantId', tenant._id))
  .paginate({{ cursor: {cursor_literal}, numItems: 2000 }});
return {{
  page: result.page.map({mapper}),
  continueCursor: result.continueCursor,
  isDone: result.isDone,
}};
"""
    return convex_run(context, query)


def convex_collect_paged(context: ToolContext, table_name: str, mapper: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    cursor: str | None = None
    while True:
        page = convex_page(context, table_name, mapper, cursor)
        rows.extend(page["page"])
        if page.get("isDone"):
            break
        cursor = page.get("continueCursor")
    return rows


def convex_snapshot(context: ToolContext) -> dict[str, list[dict[str, Any]]]:
    tenant_slug = json.dumps(context.tenant_slug)
    suppliers = convex_run(
        context,
        """
const tenant = await ctx.db.query('tenants').withIndex('by_slug', q => q.eq('slug', TENANT_SLUG_VALUE)).unique();
const suppliers = await ctx.db.query('suppliers').withIndex('by_tenant', q => q.eq('tenantId', tenant._id)).collect();
return suppliers.map((item) => ({ id: String(item._id), name: item.name }));
""".replace("TENANT_SLUG_VALUE", tenant_slug)
    )
    import_profiles = convex_run(
        context,
        """
const tenant = await ctx.db.query('tenants').withIndex('by_slug', q => q.eq('slug', TENANT_SLUG_VALUE)).unique();
const importProfiles = await ctx.db.query('importProfiles').withIndex('by_tenant', q => q.eq('tenantId', tenant._id)).collect();
return importProfiles.map((item) => ({
  id: String(item._id),
  supplierName: item.supplierName,
  name: item.name,
  filePattern: item.filePattern ?? null,
  expectedFileExtension: item.expectedFileExtension ?? null,
  supportsXlsx: item.supportsXlsx,
  supportsXls: item.supportsXls,
  status: item.status,
}));
""".replace("TENANT_SLUG_VALUE", tenant_slug)
    )

    products = convex_collect_paged(
        context,
        "products",
        """(item) => ({
    id: String(item._id),
    supplierId: item.supplierId ? String(item.supplierId) : null,
    articleNumber: item.articleNumber ?? null,
    name: item.name,
    status: item.status,
  })""",
    )
    product_prices = convex_collect_paged(
        context,
        "productPrices",
        """(item) => ({
    id: String(item._id),
    productId: String(item.productId),
    sourceKey: item.sourceKey ?? null,
    sourceFileName: item.sourceFileName ?? null,
    sourceSheetName: item.sourceSheetName ?? null,
    sourceRowNumber: item.sourceRowNumber ?? null,
    sourceColumnIndex: item.sourceColumnIndex ?? null,
    sourceColumnName: item.sourceColumnName ?? null,
    sourceValue: item.sourceValue ?? null,
    amount: item.amount,
    priceType: item.priceType,
    vatMode: item.vatMode,
  })""",
    )
    return {
        "suppliers": suppliers,
        "products": products,
        "productPrices": product_prices,
        "importProfiles": import_profiles,
    }


def parse_excel_articles(path: Path) -> set[str]:
    rows, _preview_rows, _warnings = parse_workbook(path)
    return {
        str(row["articleNumber"])
        for row in rows
        if row.get("articleNumber")
    }


def strict_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace("\u00a0", " ")
    if not text:
        return None
    text = text.replace("€", "").replace("EUR", "").replace("eur", "").strip()
    text = text.replace(" ", "")
    if not re.fullmatch(r"-?\d+(?:[.,]\d+)?|-?\d{1,3}(?:\.\d{3})+(?:,\d+)?", text):
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def profile_for_file(path: Path, profiles: list[dict[str, Any]]) -> str | None:
    extension = path.suffix.lower()
    active_profiles = [profile for profile in profiles if profile.get("status") == "active"]
    for profile in active_profiles:
        expected = profile.get("expectedFileExtension")
        if expected and expected.lower() != extension:
            continue
        if extension == ".xlsx" and not profile.get("supportsXlsx"):
            continue
        if extension == ".xls" and not profile.get("supportsXls"):
            continue
        pattern = profile.get("filePattern")
        if pattern and not fnmatch.fnmatch(path.name.lower(), pattern.lower()):
            continue
        return str(profile.get("name"))
    return None


def sampled_articles(article_numbers: set[str]) -> set[str]:
    ordered = sorted(article_numbers)
    count = len(ordered)
    if count == 0:
        return set()
    target = int(count * 0.10)
    target = min(max(target, min(5, count)), 50)
    step = max(1, count // target)
    sampled: list[str] = []
    index = step - 1
    while index < count and len(sampled) < target:
        sampled.append(ordered[index])
        index += step
    cursor = 0
    while len(sampled) < target and cursor < count:
        value = ordered[cursor]
        if value not in sampled:
            sampled.append(value)
        cursor += 1
    return set(sampled)


def raw_cell_matches(path: Path, price: dict[str, Any]) -> tuple[bool, str | None]:
    row_number = price.get("sourceRowNumber")
    column_index = price.get("sourceColumnIndex")
    sheet_name = price.get("sourceSheetName")
    if row_number is None:
        return False, "source_row_number missing"
    if column_index is None:
        return False, "source_column_index missing"
    if not sheet_name:
        return False, "source_sheet_name missing"

    analysis_path = converted_xlsx_for(path)
    if analysis_path is None:
        return False, "legacy .xls conversion missing"

    try:
        workbook = load_workbook(analysis_path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - surfaced in audit output
        return False, f"workbook open failed: {exc}"
    try:
        resolved_sheet_name = sheet_name
        if resolved_sheet_name not in workbook.sheetnames:
            resolved_sheet_name = next(
                (
                    candidate
                    for candidate in workbook.sheetnames
                    if clean_text(candidate) == clean_text(sheet_name)
                ),
                "",
            )
        if not resolved_sheet_name:
            return False, "source sheet missing"
        worksheet = workbook[resolved_sheet_name]
        if row_number < 1 or row_number > (worksheet.max_row or 0):
            return False, "source row missing"
        raw_value = worksheet.cell(row=row_number, column=int(column_index) + 1).value
        if safe_sample(raw_value) != str(price.get("sourceValue") or ""):
            return False, "matching source row/cell missing"
    finally:
        workbook.close()
    return True, None


def verdict_for(status: str, deviations: int, untraced: int, vat_open: int) -> str:
    if status in {"NO_PROFILE", "NOT_IMPORTED"} or deviations or untraced:
        return "RED"
    if status == "PARTIAL" or vat_open:
        return "YELLOW"
    return "GREEN"


def analyze_sources(context: ToolContext, snapshot: dict[str, list[dict[str, Any]]]) -> list[SourceAudit]:
    products_by_id = {product["id"]: product for product in snapshot["products"]}
    supplier_by_id = {supplier["id"]: supplier["name"] for supplier in snapshot["suppliers"]}
    prices_by_file: dict[str, list[dict[str, Any]]] = defaultdict(list)
    articles_by_file: dict[str, set[str]] = defaultdict(set)

    for price in snapshot["productPrices"]:
        file_name = price.get("sourceFileName")
        if not file_name:
            continue
        prices_by_file[file_name].append(price)
        product = products_by_id.get(price["productId"])
        article_number = product.get("articleNumber") if product else None
        if article_number:
            articles_by_file[file_name].add(str(article_number))

    audits: list[SourceAudit] = []
    for path in source_files(context):
        file_name = path.name
        profile_name = profile_for_file(path, snapshot["importProfiles"])
        excel_articles = parse_excel_articles(path)
        convex_articles = articles_by_file.get(file_name, set())
        matched_articles = excel_articles & convex_articles
        price_rows = prices_by_file.get(file_name, [])

        if excel_articles:
            coverage_pct = round((len(matched_articles) / len(excel_articles)) * 100, 1)
        else:
            coverage_pct = None

        if not price_rows and not profile_name:
            status = "NO_PROFILE"
        elif not price_rows:
            status = "NOT_IMPORTED"
        elif coverage_pct is not None and coverage_pct >= 99.5:
            status = "COVERED"
        else:
            status = "PARTIAL"

        sample = sampled_articles(convex_articles)
        deviations: list[dict[str, Any]] = []
        untraced: list[dict[str, Any]] = []
        vat_open = sum(1 for price in price_rows if price.get("vatMode") == "unknown")

        if status in {"COVERED", "PARTIAL"}:
            for price in price_rows:
                product = products_by_id.get(price["productId"], {})
                article_number = product.get("articleNumber")
                if not article_number or str(article_number) not in sample:
                    continue
                traced, reason = raw_cell_matches(path, price)
                if not traced:
                    untraced.append(
                        {
                            "file": file_name,
                            "article": article_number,
                            "product": product.get("name"),
                            "source_row": price.get("sourceRowNumber"),
                            "source_value": price.get("sourceValue"),
                            "price_type": price.get("priceType"),
                            "vat_mode": price.get("vatMode"),
                            "reason": reason,
                        }
                    )
                    continue
                source_number = strict_number(price.get("sourceValue"))
                amount = strict_number(price.get("amount"))
                if source_number is None or amount is None:
                    untraced.append(
                        {
                            "file": file_name,
                            "article": article_number,
                            "product": product.get("name"),
                            "source_row": price.get("sourceRowNumber"),
                            "source_value": price.get("sourceValue"),
                            "price_type": price.get("priceType"),
                            "vat_mode": price.get("vatMode"),
                            "reason": "numeric source value missing",
                        }
                    )
                    continue
                delta = amount - source_number
                if abs(delta) > DEVIATION_THRESHOLD:
                    deviations.append(
                        {
                            "file": file_name,
                            "article": article_number,
                            "source_row": price.get("sourceRowNumber"),
                            "source_value": price.get("sourceValue"),
                            "amount_convex": price.get("amount"),
                            "delta": delta,
                        }
                    )

        verdict = verdict_for(status, len(deviations), len(untraced), vat_open)
        audits.append(
            SourceAudit(
                path=path,
                file_name=file_name,
                profile_name=profile_name,
                excel_articles=excel_articles,
                matched_articles=matched_articles,
                coverage_pct=coverage_pct,
                status=status,
                deviations=deviations,
                untraced=untraced,
                vat_open=vat_open,
                verdict=verdict,
            )
        )

    return audits


def markdown_inventory(audits: list[SourceAudit]) -> str:
    status_order = {"COVERED": 0, "PARTIAL": 1, "NO_PROFILE": 2, "NOT_IMPORTED": 3}
    ordered = sorted(audits, key=lambda item: (status_order.get(item.status, 99), str(item.path)))
    total_deviations = sum(len(item.deviations) for item in audits)
    total_untraced = sum(len(item.untraced) for item in audits)
    seen_file_names: set[str] = set()
    total_vat_open = 0
    for item in audits:
        if item.file_name in seen_file_names:
            continue
        seen_file_names.add(item.file_name)
        total_vat_open += item.vat_open
    covered = sum(1 for item in audits if item.status == "COVERED")
    overall = "RED" if any(item.verdict == "RED" for item in audits) else ("YELLOW" if total_vat_open else "GREEN")

    lines = [
        "# Source Inventory",
        "",
        "| file | coverage% | deviations | untraced | vat_open | status |",
        "| --- | ---: | ---: | ---: | ---: | --- |",
    ]
    for item in ordered:
        rel_path = str(item.path.relative_to(ROOT)).replace("/", "\\")
        coverage = "missing" if item.coverage_pct is None else f"{item.coverage_pct:.1f}"
        lines.append(
            f"| `{rel_path}` | {coverage} | {len(item.deviations)} | {len(item.untraced)} | {item.vat_open} | {item.status} |"
        )
    lines.extend(
        [
            "",
            f"[docs/audit/02_source_inventory.md] | covered: {covered}/{len(audits)} | deviations: {total_deviations} | untraced: {total_untraced} | vat_open: {total_vat_open} | {overall}",
            "",
        ]
    )
    return "\n".join(lines)


def autosize(worksheet) -> None:
    for column_cells in worksheet.columns:
        width = 12
        for cell in column_cells:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 72))
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_workbook(audits: list[SourceAudit], snapshot: dict[str, list[dict[str, Any]]]) -> None:
    products_by_id = {product["id"]: product for product in snapshot["products"]}
    supplier_by_id = {supplier["id"]: supplier["name"] for supplier in snapshot["suppliers"]}
    profile_by_file = {audit.file_name: audit.profile_name or "NEW_PROFILE_REQUIRED" for audit in audits}

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["file", "coverage%", "deviations", "untraced", "vat_open", "verdict"])
    for item in sorted(audits, key=lambda audit: str(audit.path)):
        summary.append(
            [
                str(item.path.relative_to(ROOT)).replace("/", "\\"),
                "missing" if item.coverage_pct is None else f"{item.coverage_pct:.1f}",
                len(item.deviations),
                len(item.untraced),
                item.vat_open,
                item.verdict,
            ]
        )

    deviations = workbook.create_sheet("Deviations")
    deviations.append(["file", "article", "source_row", "source_value", "amount_convex", "delta"])
    for item in audits:
        for row in item.deviations:
            deviations.append(
                [
                    row["file"],
                    row["article"],
                    row["source_row"],
                    row["source_value"],
                    row["amount_convex"],
                    row["delta"],
                ]
            )

    untraced = workbook.create_sheet("Untraced")
    untraced.append(["file", "article", "product", "source_row", "source_value", "price_type", "vat_mode", "reason"])
    for item in audits:
        for row in item.untraced:
            untraced.append(
                [
                    row["file"],
                    row["article"],
                    row["product"],
                    row["source_row"],
                    row["source_value"],
                    row["price_type"],
                    row["vat_mode"],
                    row["reason"],
                ]
            )

    vat_open = workbook.create_sheet("Vat_Open")
    vat_open.append(["supplier", "import_profile", "file", "vat_mode", "count"])
    grouped: Counter[tuple[str, str, str, str]] = Counter()
    for price in snapshot["productPrices"]:
        if price.get("vatMode") != "unknown":
            continue
        file_name = price.get("sourceFileName") or "Onbekend bestand"
        product = products_by_id.get(price["productId"], {})
        supplier_id = product.get("supplierId")
        supplier = supplier_by_id.get(str(supplier_id), "Onbekend")
        profile = profile_by_file.get(file_name, "NEW_PROFILE_REQUIRED")
        grouped[(supplier, profile, file_name, "unknown")] += 1
    for (supplier, profile, file_name, vat_mode), count in sorted(grouped.items()):
        vat_open.append([supplier, profile, file_name, vat_mode, count])

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        worksheet.freeze_panes = "A2"
        autosize(worksheet)

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(RECONCILIATION_OUT)


def main() -> None:
    context = build_context(sys.argv[1:])
    snapshot = convex_snapshot(context)
    audits = analyze_sources(context, snapshot)

    if not context.no_write:
        AUDIT_DIR.mkdir(parents=True, exist_ok=True)
        INVENTORY_OUT.write_text(markdown_inventory(audits), encoding="utf-8")
        write_workbook(audits, snapshot)

    total_deviations = sum(len(item.deviations) for item in audits)
    total_untraced = sum(len(item.untraced) for item in audits)
    seen_file_names: set[str] = set()
    total_vat_open = 0
    for item in audits:
        if item.file_name in seen_file_names:
            continue
        seen_file_names.add(item.file_name)
        total_vat_open += item.vat_open
    green = sum(1 for item in audits if item.verdict == "GREEN")
    yellow = sum(1 for item in audits if item.verdict == "YELLOW")
    red = sum(1 for item in audits if item.verdict == "RED")
    production_status = "READY" if not (total_deviations or total_untraced or total_vat_open or red) else "BLOCKED"

    print(
        json.dumps(
            {
                "tenantSlug": context.tenant_slug,
                "target": context.target,
                "convexDeployment": context.convex_deployment,
                "convexUrl": context.convex_url,
                "envFile": str(context.env_file) if context.env_file else None,
                "envFileLoaded": context.env_file_loaded,
                "skipEnvFile": context.skip_env_file,
                "sourceFilter": context.source_filters or None,
                "writeOutputs": not context.no_write,
                "total_source_files": len(audits),
                "fully_covered_GREEN": green,
                "partial_YELLOW": yellow,
                "not_imported_RED": red,
                "total_deviations": total_deviations,
                "total_untraced": total_untraced,
                "total_vat_open": total_vat_open,
                "production_import_status": production_status,
                "threshold": DEVIATION_THRESHOLD,
                "outputs": [
                    str(INVENTORY_OUT.relative_to(ROOT)),
                    str(RECONCILIATION_OUT.relative_to(ROOT)),
                ]
                if not context.no_write
                else None,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
