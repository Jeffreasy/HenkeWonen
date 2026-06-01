from __future__ import annotations

import hashlib
import json
import math
import os
import re
import sys
import pypdf
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parent))

from audit_excel_data import (  # noqa: E402
    CODE_HEADER_RE,
    CONVERTED_DIR,
    DATA_DIR,
    ROOT,
    clean_text,
    detect_header_row,
    infer_price_type,
    infer_unit,
    infer_vat_mode,
    is_price_header as _orig_is_price_header,
    is_quantity_header,
    row_kind,
    safe_sample,
    sheet_rows,
)


def is_price_header(header: str) -> bool:
    lower = header.lower()
    if "prix" in lower or "achat" in lower or "vente" in lower:
        return True
    return _orig_is_price_header(header)


OUT_DIR = ROOT / "docs"
SUMMARY_OUT = OUT_DIR / "catalog-import-summary.md"
SUMMARY_JSON_OUT = OUT_DIR / "catalog-import-summary.json"
SAMPLE_OUT = OUT_DIR / "catalog-import-sample.md"
PREVIEW_JSON_OUT = OUT_DIR / "catalog-import-preview.json"
GENERATED_DIR = OUT_DIR / "generated"
FULL_ROWS_OUT = GENERATED_DIR / "catalog-import-preview.full.jsonl"

MAX_ROWS_PER_SHEET = 50000

DISPLAY_CATEGORIES = {
    "pvc-vloeren": "PVC Vloeren",
    # "pvc-click" is uitgesloten op verzoek van de klant (2026-05-19)
    "pvc-dryback": "PVC Dryback",
    "palletcollectie-pvc": "Palletcollectie PVC",
    "traprenovatie": "Traprenovatie",
    "tapijt": "Tapijt",
    "vinyl": "Vinyl",
    "gordijnen": "Gordijnen",
    "raambekleding": "Raambekleding",
    "wandpanelen": "Wandpanelen",
    "douchepanelen": "Douchepanelen",
    "tegels": "Tegels",
    "entreematten": "Entreematten",
    "plinten": "Plinten",
    "lijm": "Lijm",
    "kit": "Kit",
    "egaline": "Egaline",
    "ondervloer": "Ondervloer",
    "behang": "Behang",
    "roedes-railsen": "Roedes/Railsen",
    "karpetten": "Karpetten",
    "horren": "Horren",
    "verlichting": "Verlichting",
    "overig": "Overig",
}

# Categorieën die bewust worden uitgesloten van de import.
# Producten met een slug in deze set worden als 'ignored' gemarkeerd.
EXCLUDED_CATEGORY_SLUGS: set[str] = {
    "pvc-click",  # klantverzoek 2026-05-19: PVC Click niet meer tonen
}

ATTRIBUTE_HEADERS = {
    "breedte",
    "width",
    "afmeting",
    "afmetingen",
    "kamerhoog",
    "lining",
    "pattern length",
    "weight",
    "roman blinds",
    "pattern width",
    "material style",
    "washing symbols",
    "composition",
    "suitable for panel curtains",
    "full length curtains",
    "mart visser",
    "type",
    "rug",
    "projectgeschikt",
    "warmtedoorlating",
    "poolmateriaal",
    "soort backing",
    "garantie woongebruik",
    "materiaal latten",
    "materiaal backing",
    "vochtwerend",
    "click system",
    "eir",
    "class",
    "bevels",
    "bevel name",
    "structuur",
    "klasse",
    "klasse (wonen)",
    "klasse (project)",
    "klasse (industrieel)",
    "garantie wonen (jaren)",
    "garantie woongebruik",
    "garantie project (jaren)",
    "micro bevel 4v",
    "v-groef",
    "unieke planken per decor",
    "unieke planken per décor",
    "wateropname p/m2",
    "totaal gewicht in g/m2",
    "totaal hoogte in mm",
    "hoogte (mm)",
    "kg",
    "verpakking",
    "inhoud",
    "aantal kleuren",
    "diameter/ hoogte",
    "dimbaar",
    "fitting",
    "fitting/lamp",
    "kelvin/ kleur",
    "levensduur",
    "lumen",
    "merk",
    "watt",
    "wattage",
}


def normalize(value: str) -> str:
    text = value.lower().strip()
    text = text.replace("€", "eur").replace("²", "2").replace("¹", "1")
    text = re.sub(r"\s+", " ", text)
    return text


def key(value: str) -> str:
    text = normalize(value)
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def attribute_key(value: str) -> str:
    text = normalize(value)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = text.strip("_")
    return text or "attribute"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def stable_hash(parts: list[Any]) -> str:
    joined = "|".join(clean_text(part) for part in parts if clean_text(part))
    return hashlib.sha256(joined.encode("utf-8")).hexdigest()[:24]


def raw_row(headers: list[str], values: list[Any]) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for index, value in enumerate(values):
        header = headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}"
        sample = safe_sample(value)
        if sample:
            raw[header] = sample
    return raw


def preview_row(
    source_path: Path,
    analysis_path: Path,
    sheet_name: str,
    row_number: int,
    row_kind_value: str,
    headers: list[str],
    values: list[Any],
    section_label: str | None = None,
    normalized: dict[str, Any] | None = None,
    warnings: list[str] | None = None,
    errors: list[str] | None = None,
) -> dict[str, Any]:
    row_raw = raw_row(headers, values)
    status = "valid"
    if row_kind_value in {"header", "section", "empty", "ignored"}:
        status = "ignored"
    elif errors:
        status = "error"
    elif warnings:
        status = "warning"

    return {
        "rowKind": row_kind_value,
        "status": status,
        "sourceFileName": source_path.name,
        "sourceSheetName": sheet_name,
        "rowNumber": row_number,
        "rowHash": stable_hash(
            [
                str(source_path.relative_to(ROOT)),
                sheet_name,
                row_number,
                json.dumps(row_raw, ensure_ascii=False, sort_keys=True),
            ]
        ),
        "raw": row_raw,
        "normalized": normalized,
        "sectionLabel": section_label,
        "warnings": warnings or [],
        "errors": errors or [],
        "analysisPath": str(analysis_path.relative_to(ROOT)),
    }


def converted_xlsx_for(path: Path) -> Path | None:
    if path.suffix.lower() != ".xls":
        return path
    candidate = CONVERTED_DIR / f"{path.stem}.xlsx"
    return candidate if candidate.exists() else None


def code_string(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else format(value, "f").rstrip("0").rstrip(".")
    text = clean_text(value)
    return text or None


def numeric_article_number_string(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return str(int(float(value)))

    text = clean_text(value)
    if not text:
        return None
    normalized = text.replace(" ", "").replace(",", ".")
    if re.fullmatch(r"\d+(?:\.\d+)?", normalized):
        return str(int(float(normalized)))
    return text


def article_number_for(headers: list[str], values: list[Any], source_path: Path) -> str | None:
    if is_texdecor(source_path):
        return code_string(get_value(headers, values, ["Réfcom"]))
    if is_unilin(source_path):
        # Lay Red: "Material" is the SAP material number (e.g. 400105600)
        material = code_string(get_value(headers, values, ["Material"]))
        if material:
            return material
        # Moods: "SAP order code" (e.g. 398647)
        sap = code_string(get_value(headers, values, ["SAP order code"]))
        if sap:
            return sap
        return None
    value = get_value(headers, values, ["Artikelnummer", "Art.nr."])
    if source_path.name.lower() == "prijslijst traprenovatie floorlife 2025.xlsx":
        return numeric_article_number_string(value)
    return code_string(value)


def number_value(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = clean_text(value)
    if not text or text.upper() in {"#N/A", "N/A", "-"}:
        return None
    text = text.replace("€", "").replace("EUR", "").replace("eur", "")
    text = text.replace("\u00a0", " ").strip()
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def intish(value: float | None) -> int | float | None:
    if value is None:
        return None
    return int(value) if value.is_integer() else value


def headers_for(ws) -> tuple[int | None, list[str]]:
    max_col = min(ws.max_column or 80, 80)
    header_row = detect_header_row(ws, 100, max_col)
    if not header_row:
        return None, []
    header_values = next(sheet_rows(ws, header_row, header_row, max_col), [])
    headers = [clean_text(value).replace("\n", " ") for value in header_values]
    while headers and not headers[-1]:
        headers.pop()
    return header_row, headers


def get_value(headers: list[str], values: list[Any], candidates: list[str]) -> Any:
    candidate_keys = {key(candidate) for candidate in candidates}
    for index, header in enumerate(headers):
        if index >= len(values):
            continue
        if key(header) in candidate_keys:
            return values[index]
    return None


def get_text(headers: list[str], values: list[Any], candidates: list[str]) -> str | None:
    return code_string(get_value(headers, values, candidates))


def has_text(*values: str | None) -> str | None:
    parts = [value for value in values if value]
    return " ".join(parts) if parts else None


def is_ztahl(path: Path) -> bool:
    return "ztahl" in path.name.lower()


def is_ztahl_light_source_sheet(sheet_name: str) -> bool:
    return sheet_name.lower() == "lichtbronnenlijst"


def is_ztahl_skip_code(value: str | None) -> bool:
    text = (value or "").strip().lower()
    return (
        not text
        or text == "artikelnummer"
        or text.startswith("stap")
        or text.startswith("let op")
    )


def is_unilin(path: Path) -> bool:
    """True for files in the Unilin Flooring supplier directory."""
    rel = str(path).replace("\\", "/").lower()
    return "unilin flooring" in rel


def is_flexcolours(path: Path) -> bool:
    """True for FlexColours raambekleding price-matrix files (not parseable as product rows)."""
    rel = str(path).replace("\\", "/").lower()
    return "flexcolours" in rel


def is_texdecor(path: Path) -> bool:
    """True for files in the Texdecor supplier directory."""
    rel = str(path).replace("\\", "/").lower()
    return "leveranciers/texdecor" in rel or "texdecor" in path.name.lower()


def is_lamelio(path: Path) -> bool:
    """True for files in the Lamelio supplier directory."""
    rel = str(path).replace("\\", "/").lower()
    return "leveranciers/lamelio" in rel or "lamelio" in path.name.lower()


def is_hebeta(path: Path) -> bool:
    """True for files in the Hebeta Tapijt supplier directory."""
    rel = str(path).replace("\\", "/").lower()
    return "hebeta tapijt" in rel or "hebeta" in path.name.lower()


def texdecor_headers_for(ws, file_name: str) -> tuple[int | None, list[str]]:
    """Determine the header row and retrieve clean column headers for Texdecor."""
    if "CAD_CAL" in file_name:
        header_row = 2
    else:
        header_row = 3
    max_col = min(ws.max_column or 150, 150)
    header_values = next(sheet_rows(ws, header_row, header_row, max_col), [])
    headers = [clean_text(value).replace("\n", " ") for value in header_values]
    while headers and not headers[-1]:
        headers.pop()
    return header_row, headers


def ztahl_article_code(headers: list[str], values: list[Any]) -> str | None:
    return get_text(headers, values, ["Artikelnummer", "Art.nr."]) or (
        code_string(values[0]) if values else None
    )


def ztahl_import_article_code(headers: list[str], values: list[Any]) -> str | None:
    code = ztahl_article_code(headers, values)
    return code.lower() if code else None


def first_price_amount(headers: list[str], values: list[Any]) -> float | None:
    for index, header in enumerate(headers):
        if index < len(values) and is_price_header(header):
            amount = number_value(values[index])
            if amount is not None and amount > 0:
                return amount
    return None


def prepare_ztahl_headers(sheet_name: str, headers: list[str]) -> list[str]:
    if is_ztahl_light_source_sheet(sheet_name):
        return headers
    if len(headers) >= 8:
        return headers
    return [*headers, "EAN code"]


def ztahl_headers_for(ws, sheet_name: str) -> tuple[int | None, list[str]]:
    if is_ztahl_light_source_sheet(sheet_name):
        return headers_for(ws)

    max_col = min(ws.max_column or 80, 80)
    max_scan_row = min(ws.max_row or 150, 150)
    for row_number in range(1, max_scan_row + 1):
        header_values = next(sheet_rows(ws, row_number, row_number, max_col), [])
        headers = [clean_text(value).replace("\n", " ") for value in header_values]
        while headers and not headers[-1]:
            headers.pop()
        header_keys = {key(header) for header in headers if header}
        if key(headers[0] if headers else "") == "artikelnummer" and {"prijs", "model"}.issubset(header_keys):
            return row_number, prepare_ztahl_headers(sheet_name, headers)

    header_row, headers = headers_for(ws)
    return header_row, prepare_ztahl_headers(sheet_name, headers) if headers else headers


def ztahl_light_source_codes(workbook) -> set[str]:
    if "Lichtbronnenlijst" not in workbook.sheetnames:
        return set()

    ws = workbook["Lichtbronnenlijst"]
    header_row, headers = headers_for(ws)
    if not header_row or not headers:
        return set()

    codes: set[str] = set()
    max_col = len(headers)
    for values in sheet_rows(ws, header_row + 1, ws.max_row or MAX_ROWS_PER_SHEET, max_col):
        row = list(values)
        code = ztahl_article_code(headers, row)
        if is_ztahl_skip_code(code) or first_price_amount(headers, row) is None:
            continue
        codes.add(code.lower())

    return codes


def ztahl_duplicate_price_key(row: dict[str, Any]) -> tuple[Any, ...] | None:
    prices = row.get("prices")
    if not isinstance(prices, list) or not prices:
        return None

    price = prices[0]
    article_number = code_string(row.get("articleNumber"))
    if not article_number:
        return None

    return (
        row.get("sourceFileName"),
        article_number.lower(),
        price.get("priceType"),
        price.get("priceUnit"),
        price.get("amount"),
        price.get("vatMode"),
    )


def supplier_for(path: Path) -> str:
    if is_lamelio(path):
        return "Lamelio"
    name = path.name.lower()
    if "ztahl" in name:
        return "ZTAHL"
    if "headlam" in name:
        return "Headlam"
    if "interfloor" in name:
        return "Interfloor"
    if "co-pro" in name:
        return "Co-pro"
    if "ambiant" in name:
        return "Ambiant"
    if "evc" in name:
        return "EVC"
    if "vtwonen" in name or "vt wonen" in name:
        return "vtwonen"
    if is_unilin(path):
        return "Unilin Flooring"
    if "roots" in name:
        # Roots is een merk van Unilin Flooring (groothandel) — klantverzoek 2026-05-19
        return "Unilin Flooring"
    if "moduleo" in name:
        # Moduleo is een merk van Unilin Flooring (groothandel) — klantverzoek 2026-05-19
        return "Unilin Flooring"
    return "Floorlife"


def category_slug_for(path: Path, sheet_name: str, section_label: str | None, product_name: str) -> str:
    source = f"{path.name} {sheet_name}".lower()
    text = f"{source} {section_label or ''} {product_name}".lower()
    if is_lamelio(path):
        if "kit" in product_name.lower():
            return "kit"
        return "wandpanelen"
    if is_ztahl(path):
        return "verlichting"
    if is_unilin(path):
        # Lay Red en Moods zijn PVC LVT-vloercollecties van Unilin (niet click)
        # Roots-bestanden vallen hier niet onder (die staan in de HenkeWonen-map)
        return "pvc-vloeren"
    if "lijm" in source and "kit" in source and "egaline" in source:
        product_text = product_name.lower()
        if any(
            token in product_text
            for token in ["lijm", "kleefstof", "kleefstoffen", "fixeer", "antislip", "parketlijm"]
        ):
            return "lijm"
        if any(token in product_text for token in ["kit", "siliconen", "silicone", "acrylaat", "afdicht"]):
            return "kit"
        if any(
            token in product_text
            for token in ["egal", "primer", "voorstrijk", "mortel", "kwartszand", "mengemmer", "maatemmer"]
        ):
            return "egaline"
        material_text = f"{section_label or ''} {product_name}".lower()
        if any(token in material_text for token in ["kit", "siliconen", "silicone", "acrylaat", "afdicht"]):
            return "kit"
        return "egaline"
    if "douchepanelen" in source and "tegels" in source:
        detail_text = f"{section_label or ''} {product_name}".lower()
        if "tegel" in detail_text:
            return "tegels"
        return "douchepanelen"
    if "karpet" in text:
        return "karpetten"
    if "vinyl" in source:
        return "vinyl"
    if "tapijt" in source or "interfloor" in source:
        return "tapijt"
    if "palletcollectie" in source:
        return "palletcollectie-pvc"
    if "pvc" in source or "evc" in source or "roots" in source:
        if "dryback" in sheet_name.lower() or "drbyack" in sheet_name.lower():
            return "pvc-dryback"
        if "src" in sheet_name.lower() or "click" in sheet_name.lower():
            return "pvc-click"
        return "pvc-vloeren"
    if "traprenovatie" in text or "trap" in text:
        return "traprenovatie"
    if "wandpanelen" in text or "wandpaneel" in text:
        return "wandpanelen"
    if "douche" in text:
        return "douchepanelen"
    if "tegel" in text:
        return "tegels"
    if "entreemat" in text or "entreematten" in text:
        return "entreematten"
    if "plint" in text:
        return "plinten"
    if "egal" in text:
        return "egaline"
    if "kit" in text:
        return "kit"
    if "lijm" in text or "adhesive" in text:
        return "lijm"
    if "gordijn" in text or "headlam" in text or "curtain" in text:
        return "gordijnen"
    return "overig"


def product_kind_for(category_slug: str, path: Path, sheet_name: str, product_name: str) -> str:
    source = f"{path.name} {sheet_name}".lower()
    text = f"{source} {product_name}".lower()
    if category_slug == "karpetten":
        return "rug"
    if is_unilin(path):
        return "click"
    if "src" in sheet_name.lower():
        return "src"
    if "dryback" in sheet_name.lower() or "drbyack" in sheet_name.lower():
        return "dryback"
    if "click" in sheet_name.lower() or "click" in text:
        return "click"
    if category_slug == "tapijt":
        return "carpet"
    if category_slug == "vinyl":
        return "vinyl"
    if category_slug in {"wandpanelen", "douchepanelen"}:
        return "panel"
    if category_slug == "tegels":
        return "tile"
    if category_slug == "gordijnen":
        return "curtain_fabric"
    if category_slug == "entreematten":
        return "mat"
    if category_slug in {"lijm", "kit", "egaline"}:
        return "adhesive"
    if category_slug == "plinten":
        return "plinth"
    return "other"


def unit_for(category_slug: str, product_kind: str) -> str:
    if product_kind in {"curtain_fabric", "vitrage"}:
        return "m1"
    if category_slug in {"pvc-vloeren", "pvc-click", "pvc-dryback", "palletcollectie-pvc", "tapijt", "vinyl", "wandpanelen"}:
        return "m2"
    if category_slug in {"plinten", "roedes-railsen"}:
        return "meter"
    if category_slug == "traprenovatie":
        return "step"
    if category_slug in {"karpetten", "entreematten"}:
        return "piece"
    return "piece"


def product_type_for(product_kind: str) -> str:
    if product_kind in {"curtain", "fabric", "curtain_fabric", "vitrage", "roman_blind_fabric", "panel_curtain_fabric"}:
        return "made_to_measure"
    return "standard"


def commercial_names(headers: list[str], values: list[Any]) -> list[dict[str, str]]:
    names = []
    for brand, collection_header, color_header in [
        ("Ambiant", "Ambiant Collectie", "Ambiant Kleur"),
        ("Floorlife", "Floorlife Collectie", "Floorlife Kleur"),
    ]:
        collection = get_text(headers, values, [collection_header])
        color = get_text(headers, values, [color_header])
        display = has_text(brand, collection, color)
        if collection or color:
            item: dict[str, str] = {"brandName": brand, "displayName": display or brand}
            if collection:
                item["collectionName"] = collection
            if color:
                item["colorName"] = color
            names.append(item)
    return names


def product_name_for(headers: list[str], values: list[Any], path: Path, sheet_name: str) -> str | None:
    if is_lamelio(path):
        return get_text(headers, values, ["NL_Title_Short"])
    if is_texdecor(path):
        collection = get_text(headers, values, ["Collection Réfcom"])
        dessin = get_text(headers, values, ["Nom Dessin commercial"])
        return has_text(collection, dessin) or get_text(headers, values, ["Nom produit"]) or get_text(headers, values, ["Nom réfcom"])
    supplier = supplier_for(path)
    if supplier == "ZTAHL":
        article_number = ztahl_article_code(headers, values)
        if is_ztahl_skip_code(article_number) or first_price_amount(headers, values) is None:
            return None
        if is_ztahl_light_source_sheet(sheet_name):
            return has_text(
                get_text(headers, values, ["Merk"]),
                get_text(headers, values, ["Omschrijving"]),
            ) or get_text(headers, values, ["Omschrijving"])
        return has_text(
            get_text(headers, values, ["Model"]),
            get_text(headers, values, ["Uitvoering"]),
        ) or get_text(headers, values, ["Model"])
    if "entreematten" in path.name.lower():
        return code_string(values[1]) if len(values) > 1 else None
    if supplier == "Headlam":
        return has_text(
            get_text(headers, values, ["Quality"]),
            get_text(headers, values, ["Design"]),
            get_text(headers, values, ["Type"]),
        )
    if supplier == "Unilin Flooring":
        if "roots" in path.name.lower():
            # Roots collectie: eigen kolomnamen (Material Description / Decor name)
            return has_text(
                get_text(headers, values, ["Material Description"]),
                get_text(headers, values, ["Decor name"]),
                get_text(headers, values, ["Commercial Code"]),
            )
        # Lay Red: full material description (e.g. "LAYRED COUNTRY OAK 54991")
        mat_descr = get_text(headers, values, ["Material Descr.", "Material Descr"])
        if mat_descr:
            return mat_descr
        # Moods: pattern name + code (e.g. "Rectangle Mono 647")
        return get_text(headers, values, ["Pattern name + code", "Pattern name"])
    if supplier == "Roots":  # legacy — niet meer bereikbaar na 2026-05-19 fix
        return has_text(
            get_text(headers, values, ["Material Description"]),
            get_text(headers, values, ["Decor name"]),
            get_text(headers, values, ["Commercial Code"]),
        )
    aliases = commercial_names(headers, values)
    if aliases:
        return aliases[-1]["displayName"]
    return has_text(
        get_text(headers, values, ["Kwaliteitsnaam"]),
        get_text(headers, values, ["Kwaliteit"]),
        get_text(headers, values, ["Omschrijving"]),
        get_text(headers, values, ["Material Description"]),
        get_text(headers, values, ["Decor name"]),
        get_text(headers, values, ["Afmeting (cm)"]),
        sheet_name if supplier == "Interfloor" else None,
    )


def set_dimension_fields(row: dict[str, Any], headers: list[str], values: list[Any], source_path: Path) -> None:
    if is_texdecor(source_path):
        width_cm = number_value(get_value(headers, values, ["Laize totale Cm", "Laize utilie Cm"]))
        length_ml = number_value(get_value(headers, values, ["Longuer pièce Ml"]))
        if width_cm is not None:
            row["widthMm"] = width_cm * 10
        if length_ml is not None:
            row["lengthMm"] = length_ml * 1000
    elif is_lamelio(source_path):
        desc = get_text(headers, values, ["NL_Description_Short", "NL_Description_Long"])
        width_mm, length_mm = None, None
        if desc:
            length_match = re.search(r"lengte\s+[^.]*?\b(\d+(?:[.,]\d+)?)\s*cm", desc, re.IGNORECASE)
            if length_match:
                length_mm = float(length_match.group(1).replace(",", ".")) * 10
            width_match = re.search(r"breedte\s+[^.]*?\b(\d+(?:[.,]\d+)?)\s*cm", desc, re.IGNORECASE)
            if width_match:
                width_mm = float(width_match.group(1).replace(",", ".")) * 10

        title = get_text(headers, values, ["NL_Title_Short"]) or ""
        if length_mm is None:
            title_length_match = re.search(r"\b(\d+)\s*cm\b", title)
            if title_length_match:
                length_mm = float(title_length_match.group(1)) * 10
            elif "wandpaneel" in title.lower() or "strip" in title.lower():
                length_mm = 2700.0

        if width_mm is None and ("wandpaneel" in title.lower() or "3d wandpaneel" in title.lower() or "strip" in title.lower()):
            width_mm = 122.0

        if length_mm is not None:
            row["lengthMm"] = intish(length_mm)
        if width_mm is not None:
            row["widthMm"] = intish(width_mm)
    else:
        width = number_value(get_value(headers, values, ["Breedte (cm)", "Breedte", "Breedte in cm", "Width", "W"]))
        length = number_value(get_value(headers, values, ["Lengte (cm)", "Lengte in cm", "Lengte plank (cm)", "Lengte mm", "Lengte", "Length", "L"]))
        if width is not None:
            row["widthMm"] = width * 10 if width < 1000 else width
        if length is not None:
            row["lengthMm"] = length * 10 if length < 1000 else length

    thickness = number_value(get_value(headers, values, ["Dikte (mm)", "Totale dikte (mm)", "Totaal dikte (mm)", "Dikte mm", "Dikte in mm", "Dikte", "Tick"]))
    wear = number_value(get_value(headers, values, ["Dikte toplaag (mm)", "Toplaag in mm", "Toplaag (mm)", "Toplaag mm", "Toplaag"]))
    package_m2 = number_value(get_value(headers, values, ["Aantal m2 per pak", "Pakinhoud (m²)", "Pakinhoud", "m2"]))
    pieces = number_value(get_value(headers, values, ["Aantal panelen per pak", "Planken per pak", "Panels"]))
    packs = number_value(get_value(headers, values, ["Aantal pakken per pallet", "Aantal pakker per pallet", "Pakken per pallet", "Packs"]))

    if thickness is not None:
        row["thicknessMm"] = thickness
    if wear is not None:
        row["wearLayerMm"] = wear
    if package_m2 is not None:
        row["packageContentM2"] = package_m2
    if pieces is not None:
        row["piecesPerPackage"] = intish(pieces)
    if packs is not None:
        row["packagesPerPallet"] = intish(packs)


def set_quantity_fields(row: dict[str, Any], headers: list[str], values: list[Any]) -> None:
    for index, header in enumerate(headers):
        if index >= len(values) or not is_quantity_header(header):
            continue
        number = number_value(values[index])
        lower = normalize(header)
        text = code_string(values[index])
        if "trailer" in lower and number is not None:
            row["trailerQuantity"] = intish(number)
        elif "pallet" in lower and number is not None:
            row["palletQuantity"] = intish(number)
        elif "bundel" in lower and number is not None:
            row["bundleSize"] = intish(number)
        elif "planken per pak" in lower and number is not None:
            row["piecesPerPackage"] = intish(number)
        elif "pakinhoud" in lower and number is not None:
            row["packageContentM2"] = number
        elif "besteleenheid" in lower or "afname" in lower:
            row["orderUnit"] = text
        elif "m2 per pak" in lower or "m² per pak" in lower:
            row["packageContentM2"] = number


def valid_from_for(path: Path, header: str | None = None) -> int | None:
    text = f"{path.name} {header or ''}".lower()
    match = re.search(r"(?:vanaf\s*)?(\d{1,2})[./-](\d{1,2})[./-](\d{4})", text)
    if not match:
        return None
    day, month, year = (int(part) for part in match.groups())
    try:
        from datetime import datetime, timezone

        return int(datetime(year, month, day, tzinfo=timezone.utc).timestamp() * 1000)
    except ValueError:
        return None


def year_for(path: Path) -> int | None:
    years = [int(value) for value in re.findall(r"20\d{2}", path.name)]
    return max(years) if years else None


def attributes_for(headers: list[str], values: list[Any]) -> dict[str, Any] | None:
    attributes: dict[str, Any] = {}
    for index, header in enumerate(headers):
        if index >= len(values) or not header:
            continue
        if is_price_header(header) or is_quantity_header(header) or CODE_HEADER_RE.search(header):
            continue
        normalized = normalize(header)
        if normalized in ATTRIBUTE_HEADERS or any(token in normalized for token in ATTRIBUTE_HEADERS):
            value = code_string(values[index])
            if value:
                attributes[attribute_key(header)] = value
    return attributes or None


def price_unit_for(header: str, category_slug: str) -> str:
    if category_slug == "behang":
        return "roll"
    if category_slug == "gordijnen":
        return "m1"
    if category_slug == "wandpanelen":
        return "piece"
    unit = infer_unit(header)
    if unit != "custom":
        return unit
    lower = normalize(header)
    if "adviesverkoopprijs vanaf" in lower:
        return "pack"
    if category_slug == "gordijnen" and "consumer" in lower:
        return "m1"
    if category_slug == "verlichting":
        return "piece"
    if category_slug in {"karpetten", "douchepanelen", "wandpanelen"}:
        return "piece"
    return unit


def price_type_for(header: str, source_path: Path) -> str:
    if is_ztahl(source_path):
        lower = source_path.name.lower()
        if "inkoop" in lower:
            return "purchase"
        if "verkoop" in lower:
            return "advice_retail"
    if is_texdecor(source_path):
        lower = header.lower()
        if "achat" in lower or "vente ned" in lower:
            return "purchase"
        if "public" in lower:
            return "advice_retail"
    return infer_price_type(header)


def vat_mode_for(header: str, source_path: Path) -> str:
    if is_ztahl(source_path):
        lower = source_path.name.lower()
        if "inkoop" in lower:
            return "exclusive"
        if "verkoop" in lower:
            return "inclusive"
    if is_texdecor(source_path):
        lower = header.lower()
        if "achat" in lower or "vente ned" in lower:
            return "exclusive"
        if "public" in lower:
            return "inclusive"
    return infer_vat_mode(header)


def source_key_for_price(
    row: dict[str, Any],
    source_path: Path,
    row_number: int,
    column_index: int,
    header: str,
    price_type: str,
    price_unit: str,
    amount: float,
    vat_mode: str,
) -> str:
    if is_ztahl(source_path):
        return stable_hash(
            [
                row["importKey"],
                row.get("fileHash") or row["sourcePath"],
                row["sourceSheetName"],
                row_number,
                column_index,
                price_type,
                price_unit,
                round(amount, 4),
                vat_mode,
                header,
            ]
        )

    return stable_hash(
        [
            row["importKey"],
            row.get("fileHash") or row["sourcePath"],
            row["sourceSheetName"],
            row_number,
            column_index,
            header,
            price_type,
            price_unit,
        ]
    )


def prices_for(
    headers: list[str],
    values: list[Any],
    row: dict[str, Any],
    source_path: Path,
) -> list[dict[str, Any]]:
    if is_lamelio(source_path):
        msrp = number_value(get_value(headers, values, ["Price"]))
        if msrp is not None and msrp > 0:
            prices = []
            price_index = headers.index("Price") if "Price" in headers else 8

            # 1. purchase price (excl VAT)
            purchase_amount = msrp / 1.21 * 0.60
            purchase_key = stable_hash([
                row["importKey"],
                row.get("fileHash") or row["sourcePath"],
                row["sourceSheetName"],
                row["sourceRowNumber"],
                price_index,
                "purchase",
                "piece",
                round(purchase_amount, 4),
                "exclusive",
                "Price"
            ])
            prices.append({
                "sourceKey": purchase_key,
                "priceType": "purchase",
                "priceUnit": "piece",
                "amount": round(purchase_amount, 4),
                "vatRate": 21,
                "vatMode": "exclusive",
                "validFrom": row.get("validFrom"),
                "currency": "EUR",
                "sourceColumnName": "Price",
                "sourceColumnIndex": price_index,
                "sourceValue": safe_sample(get_value(headers, values, ["Price"])),
            })

            # 2. advice retail price (incl VAT)
            retail_key = stable_hash([
                row["importKey"],
                row.get("fileHash") or row["sourcePath"],
                row["sourceSheetName"],
                row["sourceRowNumber"],
                price_index,
                "advice_retail",
                "piece",
                round(msrp, 4),
                "inclusive",
                "Price"
            ])
            prices.append({
                "sourceKey": retail_key,
                "priceType": "advice_retail",
                "priceUnit": "piece",
                "amount": round(msrp, 4),
                "vatRate": 21,
                "vatMode": "inclusive",
                "validFrom": row.get("validFrom"),
                "currency": "EUR",
                "sourceColumnName": "Price",
                "sourceColumnIndex": price_index,
                "sourceValue": safe_sample(get_value(headers, values, ["Price"])),
            })
            return prices
        return []

    prices = []
    for index, header in enumerate(headers):
        if index >= len(values) or not header or not is_price_header(header):
            continue
        amount = number_value(values[index])
        if amount is None or amount <= 0:
            continue
        price_type = price_type_for(header, source_path)
        price_unit = price_unit_for(header, row["categorySlug"])
        vat_mode = vat_mode_for(header, source_path)
        source_key = source_key_for_price(
            row,
            source_path,
            row["sourceRowNumber"],
            index,
            header,
            price_type,
            price_unit,
            amount,
            vat_mode,
        )
        prices.append(
            {
                "sourceKey": source_key,
                "priceType": price_type,
                "priceUnit": price_unit,
                "amount": round(amount, 4),
                "vatRate": 21,
                "vatMode": vat_mode,
                "validFrom": valid_from_for(source_path, header) or row.get("validFrom"),
                "currency": "EUR",
                "sourceColumnName": header,
                "sourceColumnIndex": index,
                "sourceValue": safe_sample(values[index]),
            }
        )
    return prices


def build_row(
    source_path: Path,
    analysis_path: Path,
    sheet_name: str,
    headers: list[str],
    values: list[Any],
    row_number: int,
    section_label: str | None,
    file_hash: str,
) -> dict[str, Any] | None:
    product_name = product_name_for(headers, values, source_path, sheet_name)
    if not product_name:
        return None

    if is_texdecor(source_path):
        brand_code = clean_text(values[0]) if values else None
        if brand_code == "CAD":
            supplier_name = "Casadeco"
        elif brand_code == "CAL":
            supplier_name = "Caselio"
        elif brand_code in ("CAS", "CAM"):
            supplier_name = "Casamance"
        else:
            supplier_name = "Texdecor"

        support_type = clean_text(get_value(headers, values, ["Nom Type support"]))
        if support_type in ("Papier  peint", "Frise", "Stickers"):
            category_slug = "behang"
            product_kind = "wallpaper"
            unit = "roll"
        elif support_type in ("Tissus", "Panoramique Tissu"):
            category_slug = "gordijnen"
            product_kind = "curtain_fabric"
            unit = "m1"
        elif support_type in ("Panoramique Papier Peint", "Panoramique Revêtement", "Revêtement", "Affiche"):
            category_slug = "wandpanelen"
            product_kind = "panel"
            unit = "piece"
        else:
            category_slug = "overig"
            product_kind = "other"
            unit = "piece"

        category_name = DISPLAY_CATEGORIES.get(category_slug, "Overig")
    else:
        supplier_name = supplier_for(source_path)
        category_slug = category_slug_for(source_path, sheet_name, section_label, product_name)

    # Sla uitgesloten categorieën over (EXCLUDED_CATEGORY_SLUGS)
    if category_slug in EXCLUDED_CATEGORY_SLUGS:
        return None

    if not is_texdecor(source_path):
        category_name = DISPLAY_CATEGORIES.get(category_slug, "Overig")
        product_kind = product_kind_for(category_slug, source_path, sheet_name, product_name)
        unit = unit_for(category_slug, product_kind)
        if is_lamelio(source_path):
            unit = "piece"

    aliases = commercial_names(headers, values)
    article_number = article_number_for(headers, values, source_path)
    if "entreematten" in source_path.name.lower() and not article_number:
        article_number = code_string(values[0]) if values else None
    supplier_code = get_text(headers, values, ["Supplier Code", "SAP codes floors"])
    commercial_code = get_text(headers, values, ["Commercial Code", "Commercial code"])
    ean = get_text(headers, values, ["EAN", "EAN code", "EAN-code", "EAN Code", "EAN/UPC", "Code barre"])
    color = get_text(
        headers,
        values,
        ["Kleur", "Kleurindicatie", "Kleurnummer", "Design", "Decor name", "Uitvoering", "Kelvin/ kleur", "Color", "Nom couleur"],
    )
    supplier_group = get_text(headers, values, ["Artikelgroep", "Soort"])
    if is_ztahl(source_path):
        article_number = ztahl_import_article_code(headers, values)
        if is_ztahl_light_source_sheet(sheet_name):
            collection_name = "Lichtbronnen"
        else:
            collection_name = section_label
        brand_name = get_text(headers, values, ["Merk"]) or supplier_name
    elif "entreematten" in source_path.name.lower():
        collection_name = section_label
        brand_name = supplier_name
    elif is_unilin(source_path):
        # Lay Red / Moods — Unilin PVC-vloercollecties
        collection_name = (
            get_text(headers, values, ["Pattern name"])
            or section_label
        )
        brand_name = supplier_name  # "Unilin Flooring"
    elif "roots" in source_path.name.lower():
        # Roots is een Unilin-merk; sla op als merk zodat het zichtbaar blijft
        collection_name = section_label
        brand_name = "Roots"
    elif is_texdecor(source_path):
        collection_name = get_text(headers, values, ["Collection Réfcom"]) or section_label
        brand_name = supplier_name
    elif is_lamelio(source_path):
        collection_name = None
        for col_name in ["Vasco", "Olmo", "Milo", "Asti", "Amber", "Onda", "Infinity", "Allure"]:
            if col_name.lower() in product_name.lower():
                collection_name = col_name
                break
        if not collection_name:
            collection_name = "Overig"
        brand_name = "Lamelio"
    else:
        collection_name = (
            get_text(headers, values, ["Quality"])
            or get_text(headers, values, ["Kwaliteit"])
            or get_text(headers, values, ["Material Description"])
            or section_label
        )
        brand_name = get_text(headers, values, ["Company"]) or supplier_name
    source_rel = str(source_path.relative_to(ROOT))
    if is_lamelio(source_path):
        import_identity = f"{ean or ''}_{product_name}"
    else:
        import_identity = article_number or supplier_code or commercial_code or ean or product_name
    import_key = stable_hash([supplier_name, category_name, import_identity])

    row: dict[str, Any] = {
        "importKey": import_key,
        "sourceFileName": source_path.name,
        "sourceSheetName": sheet_name,
        "sourcePath": source_rel,
        "analysisPath": str(analysis_path.relative_to(ROOT)),
        "fileHash": file_hash,
        "sourceRowNumber": row_number,
        "year": year_for(source_path),
        "validFrom": valid_from_for(source_path),
        "supplierName": supplier_name,
        "brandName": brand_name,
        "collectionName": collection_name,
        "sectionLabel": section_label,
        "categorySlug": category_slug,
        "categoryName": category_name,
        "productName": product_name,
        "colorName": color,
        "articleNumber": article_number,
        "supplierCode": supplier_code,
        "commercialCode": commercial_code,
        "supplierProductGroup": supplier_group,
        "ean": ean,
        "productKind": product_kind,
        "productType": product_type_for(product_kind),
        "commercialNames": aliases or None,
        "unit": unit,
    }
    set_dimension_fields(row, headers, values, source_path)
    set_quantity_fields(row, headers, values)
    attrs = attributes_for(headers, values)
    if attrs:
        row["attributes"] = attrs
    row["prices"] = prices_for(headers, values, row, source_path)
    return {key_: value for key_, value in row.items() if value is not None}


def zero_or_no_price_warning(warnings: list[str]) -> bool:
    return any("zonder bruikbare prijs" in warning.lower() for warning in warnings)


ROW_KIND_LABELS = {
    "header": "Kopregel",
    "section": "Sectieregel",
    "product": "Productregel",
    "empty": "Lege regel",
    "warning": "Waarschuwing",
    "error": "Fout",
    "ignored": "Genegeerde regel",
}

STATUS_LABELS = {
    "valid": "Geldig",
    "warning": "Waarschuwing",
    "error": "Fout",
    "ignored": "Genegeerd",
    "imported": "Geïmporteerd",
}


def label_row_kind(value: str) -> str:
    return ROW_KIND_LABELS.get(value, value)


def label_status(value: str) -> str:
    return STATUS_LABELS.get(value, value)


def parse_hebeta_line(line: str, filename: str) -> dict[str, Any] | None:
    line = line.strip()
    if not line:
        return None
    # Skip header and footer lines
    if (line.startswith("Versie:") or line.startswith("Kwaliteit ") or
        line.startswith("Advies") or line.startswith("Levering:") or
        line.startswith("Tijdzending") or line.startswith("Voor het ") or
        line.startswith("Levering 12.00") or line.startswith("Tapijtcollectie") or
        line.startswith("HEBETA |") or line.startswith("MONTINIQUE |") or
        "orderwaarde" in line or "vervoerder" in line):
        return None
    # Also skip any time/price listing lines
    if (re.search(r"Levering \d{2}\.\d{2} uur", line) or
        re.search(r"afwijkingsmarge", line) or
        re.search(r"Schade:", line) or
        re.search(r"pakbon", line)):
        return None

    project_indicator = None
    if "✓" in line:
        project_indicator = "✓"
    elif "" in line:
        project_indicator = ""
    else:
        return None  # Not a product row

    left, right = line.split(project_indicator, 1)
    left = left.strip()
    right = right.strip()

    quality = None
    for q in ["Vario Tapijttegel", "Projecta Tapijttegel", "Marble Fushion"]:
        if left.startswith(q):
            quality = q
            break
    if not quality:
        parts = left.split(" ", 1)
        quality = parts[0]

    samenstelling_and_weight = left[len(quality):].strip()
    weight_match = re.search(r"\b(\d+)\s*gram\b", samenstelling_and_weight)
    weight = None
    if weight_match:
        weight = weight_match.group(0)
        samenstelling = samenstelling_and_weight[:weight_match.start()].strip()
    else:
        samenstelling = samenstelling_and_weight

    right_parts = right.split(" ", 1)
    breedte = right_parts[0]
    remainder = right_parts[1].strip()

    garantie_match = re.match(r"^(\d+)\s*jaar", remainder)
    garantie = None
    if garantie_match:
        garantie = garantie_match.group(0)
        price_part = remainder[garantie_match.end():].strip()
    else:
        price_part = remainder

    prices = re.findall(r"€\s*(\d+(?:[.,]\d+)?)(?:\s*p/m²)?", price_part)
    purchase_price = None
    retail_price = None
    if "Montinique" in filename:
        if len(prices) >= 1:
            retail_price = float(prices[0].replace(",", "."))
    else:
        if len(prices) >= 2:
            purchase_price = float(prices[0].replace(",", "."))
            retail_price = float(prices[1].replace(",", "."))
        elif len(prices) == 1:
            retail_price = float(prices[0].replace(",", "."))

    if purchase_price is None and retail_price is not None:
        lookup = {
            39.95: 16.00, 57.95: 23.20, 59.00: 23.60, 79.00: 31.60, 89.95: 36.00,
            119.95: 48.00, 129.00: 51.60, 149.00: 59.60, 159.00: 63.60, 169.00: 67.60,
            179.00: 71.60, 189.00: 75.60, 219.00: 87.60, 269.00: 107.60, 270.25: 108.10,
            329.00: 131.60, 339.00: 135.60, 519.00: 235.95, 579.00: 263.20, 609.00: 276.85,
            699.00: 317.75, 759.00: 345.00, 779.00: 354.10, 28.95: 18.95, 29.95: 19.95
        }
        purchase_price = lookup.get(retail_price)
        if purchase_price is None:
            if retail_price >= 500:
                purchase_price = round(retail_price * 0.4546, 2)
            else:
                purchase_price = round(retail_price * 0.40, 2)

    return {
        "quality": quality,
        "samenstelling": samenstelling,
        "weight": weight,
        "project": project_indicator == "✓",
        "breedte": breedte,
        "garantie": garantie,
        "purchase_price": purchase_price,
        "retail_price": retail_price,
    }


def parse_pdf(source_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    preview_rows: list[dict[str, Any]] = []
    warnings: list[str] = []

    file_hash = sha256(source_path)
    filename = source_path.name

    headers = [
        "Kwaliteit",
        "Samenstelling",
        "Gewicht",
        "Projectgeschikt",
        "Breedte",
        "Garantie woongebruik",
        "Inkoopprijs",
        "Adviesverkoopprijs"
    ]

    reader = pypdf.PdfReader(source_path)
    row_number = 0
    for page in reader.pages:
        text = page.extract_text()
        if not text:
            continue
        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue

            if (line.startswith("Versie:") or line.startswith("Kwaliteit ") or
                line.startswith("Advies") or line.startswith("Levering:") or
                line.startswith("Tijdzending") or line.startswith("Voor het ") or
                line.startswith("Levering 12.00") or line.startswith("Tapijtcollectie") or
                line.startswith("HEBETA |") or line.startswith("MONTINIQUE |") or
                "orderwaarde" in line or "vervoerder" in line or
                re.search(r"Levering \d{2}\.\d{2} uur", line) or
                re.search(r"afwijkingsmarge", line) or
                re.search(r"Schade:", line) or
                re.search(r"pakbon", line)):
                continue

            parsed = parse_hebeta_line(line, filename)
            row_number += 1

            if not parsed:
                preview_rows.append(
                    preview_row(
                        source_path=source_path,
                        analysis_path=source_path,
                        sheet_name="PDF",
                        row_number=row_number,
                        row_kind_value="ignored",
                        headers=["Regel"],
                        values=[line],
                        section_label=None,
                        normalized=None,
                        warnings=["Rij kon niet naar een productregel worden genormaliseerd."],
                    )
                )
                continue

            quality = parsed["quality"]
            samenstelling = parsed["samenstelling"]
            weight = parsed["weight"]
            project = parsed["project"]
            breedte = parsed["breedte"]
            garantie = parsed["garantie"]
            purchase_price = parsed["purchase_price"]
            retail_price = parsed["retail_price"]

            supplier_name = "Hebeta"
            brand_name = "Montinique" if "Montinique" in filename else "Hebeta"
            category_slug = "tapijt"
            category_name = "Tapijt"

            width_mm = None
            length_mm = None
            if breedte == "50x50":
                width_mm = 500
                length_mm = 500
                product_kind = "tile"
                unit = "m2"
            else:
                try:
                    width_mm = int(breedte) * 10
                except ValueError:
                    width_mm = 4000
                product_kind = "carpet"
                unit = "m2"

            import_key = stable_hash([supplier_name, category_name, quality])

            row: dict[str, Any] = {
                "importKey": import_key,
                "sourceFileName": filename,
                "sourceSheetName": "PDF",
                "sourcePath": str(source_path.relative_to(ROOT)),
                "analysisPath": str(source_path.relative_to(ROOT)),
                "fileHash": file_hash,
                "sourceRowNumber": row_number,
                "year": 2026,
                "validFrom": "2026-01-01",
                "supplierName": supplier_name,
                "brandName": brand_name,
                "collectionName": quality,
                "categorySlug": category_slug,
                "categoryName": category_name,
                "productName": quality,
                "productKind": product_kind,
                "productType": "standard",
                "unit": unit,
                "widthMm": width_mm,
                "lengthMm": length_mm,
                "attributes": {
                    "samenstelling": samenstelling,
                    "poolgewicht": weight,
                    "projectgeschikt": "x" if project else None,
                    "garantie_woongebruik": garantie
                }
            }

            row["attributes"] = {k: v for k, v in row["attributes"].items() if v is not None}
            if not row["attributes"]:
                del row["attributes"]

            prices: list[dict[str, Any]] = []

            def add_price(price_type: str, vat_mode: str, base_amount: float, price_unit: str, divisor: float = 1.0):
                amount = round(base_amount / divisor, 4)
                price_key = stable_hash([
                    import_key,
                    file_hash,
                    price_type,
                    price_unit,
                    amount,
                    vat_mode,
                    "Prijslijst"
                ])
                prices.append({
                    "sourceKey": price_key,
                    "priceType": price_type,
                    "priceUnit": price_unit,
                    "amount": amount,
                    "vatRate": 21,
                    "vatMode": vat_mode,
                    "validFrom": "2026-01-01",
                    "currency": "EUR",
                    "sourceColumnName": "Prijs",
                    "sourceColumnIndex": 6 if price_type == "purchase" else 7,
                    "sourceValue": f"€ {base_amount:.2f}",
                })

            if breedte == "50x50":
                if purchase_price is not None:
                    add_price("purchase", "exclusive", purchase_price, "m2")
                if retail_price is not None:
                    add_price("advice_retail", "inclusive", retail_price, "m2")
            else:
                if purchase_price is not None:
                    add_price("purchase", "exclusive", purchase_price, "m1")
                    add_price("purchase", "exclusive", purchase_price, "m2", divisor=4.0)
                if retail_price is not None:
                    add_price("advice_retail", "inclusive", retail_price, "m1")
                    add_price("advice_retail", "inclusive", retail_price, "m2", divisor=4.0)

            row["prices"] = prices
            rows.append(row)

            row_values = [
                quality,
                samenstelling,
                weight,
                "✓" if project else "",
                breedte,
                garantie,
                f"€ {purchase_price:.2f}" if purchase_price is not None else None,
                f"€ {retail_price:.2f}" if retail_price is not None else None
            ]

            preview_rows.append(
                preview_row(
                    source_path=source_path,
                    analysis_path=source_path,
                    sheet_name="PDF",
                    row_number=row_number,
                    row_kind_value="product",
                    headers=headers,
                    values=row_values,
                    section_label=None,
                    normalized=row,
                    warnings=None,
                )
            )

    return rows, preview_rows, warnings


def parse_workbook(source_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    if source_path.suffix.lower() == ".pdf":
        return parse_pdf(source_path)

    analysis_path = converted_xlsx_for(source_path)
    if analysis_path is None:
        return [], [], [f"{source_path.name}: legacy .xls is not converted in .audit-cache"]

    file_hash = sha256(source_path)
    rows: list[dict[str, Any]] = []
    preview_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    workbook = load_workbook(analysis_path, read_only=True, data_only=True)
    try:
        is_ztahl_file = is_ztahl(source_path)
        is_texdecor_file = is_texdecor(source_path)
        light_source_codes = ztahl_light_source_codes(workbook) if is_ztahl_file else set()
        seen_ztahl_price_keys: set[tuple[Any, ...]] = set()
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            if is_ztahl_file:
                header_row, headers = ztahl_headers_for(ws, sheet_name)
            elif is_texdecor_file:
                header_row, headers = texdecor_headers_for(ws, source_path.name)
            else:
                header_row, headers = headers_for(ws)
            if not header_row or not headers:
                continue
            max_col = len(headers)
            preview_rows.append(
                preview_row(
                    source_path,
                    analysis_path,
                    sheet_name,
                    header_row,
                    "header",
                    headers,
                    headers,
                )
            )
            code_indexes = [
                index for index, header in enumerate(headers) if CODE_HEADER_RE.search(header)
            ]
            price_indexes = [
                index for index, header in enumerate(headers) if is_price_header(header)
            ]
            section_label: str | None = None
            empty_streak = 0
            start = header_row + 1
            max_row = min(ws.max_row or MAX_ROWS_PER_SHEET, MAX_ROWS_PER_SHEET)
            for row_number, row_values in enumerate(
                sheet_rows(ws, start, max_row, max_col),
                start=start,
            ):
                values = list(row_values)
                if is_texdecor_file and values and values[0] == "BRAND":
                    continue
                kind = row_kind(values, code_indexes, price_indexes)
                if kind == "empty":
                    empty_streak += 1
                    preview_rows.append(
                        preview_row(
                            source_path,
                            analysis_path,
                            sheet_name,
                            row_number,
                            "empty",
                            headers,
                            values,
                            section_label,
                        )
                    )
                    if row_number > start + 100 and empty_streak >= 500:
                        break
                    continue
                empty_streak = 0
                if kind == "section":
                    section_label = clean_text(next((value for value in values if clean_text(value)), ""))
                    preview_rows.append(
                        preview_row(
                            source_path,
                            analysis_path,
                            sheet_name,
                            row_number,
                            "section",
                            headers,
                            values,
                            section_label,
                        )
                    )
                    continue
                normalized = build_row(
                    source_path,
                    analysis_path,
                    sheet_name,
                    headers,
                    values,
                    row_number,
                    section_label,
                    file_hash,
                )
                if normalized:
                    has_identity = any(
                        normalized.get(field)
                        for field in ["articleNumber", "supplierCode", "commercialCode", "ean"]
                    )
                    if not normalized.get("prices"):
                        if not has_identity:
                            section_label = normalized["productName"]
                            preview_rows.append(
                                preview_row(
                                    source_path,
                                    analysis_path,
                                    sheet_name,
                                    row_number,
                                    "section",
                                    headers,
                                    values,
                                    section_label,
                                    normalized,
                                )
                            )
                        else:
                            preview_rows.append(
                                preview_row(
                                    source_path,
                                    analysis_path,
                                    sheet_name,
                                    row_number,
                                    "ignored",
                                    headers,
                                    values,
                                    section_label,
                                    normalized,
                                    ["Productregel zonder bruikbare prijs is overgeslagen."],
                                )
                        )
                        continue
                    if is_ztahl_file:
                        article_code = code_string(normalized.get("articleNumber"))
                        if (
                            article_code
                            and not is_ztahl_light_source_sheet(sheet_name)
                            and article_code.lower() in light_source_codes
                        ):
                            preview_rows.append(
                                preview_row(
                                    source_path,
                                    analysis_path,
                                    sheet_name,
                                    row_number,
                                    "ignored",
                                    headers,
                                    values,
                                    section_label,
                                    normalized,
                                    [
                                        "ZTAHL lichtbron staat ook in Lichtbronnenlijst en is daar leidend; assortimentverwijzing overgeslagen."
                                    ],
                                )
                            )
                            continue
                        duplicate_key = ztahl_duplicate_price_key(normalized)
                        if duplicate_key and duplicate_key in seen_ztahl_price_keys:
                            preview_rows.append(
                                preview_row(
                                    source_path,
                                    analysis_path,
                                    sheet_name,
                                    row_number,
                                    "ignored",
                                    headers,
                                    values,
                                    section_label,
                                    normalized,
                                    [
                                        "Exacte ZTAHL prijsregel eerder gezien; dubbele assortimentvermelding overgeslagen."
                                    ],
                                )
                            )
                            continue
                        if duplicate_key:
                            seen_ztahl_price_keys.add(duplicate_key)
                    row_warnings = []
                    if any(price.get("vatMode") == "unknown" for price in normalized.get("prices", [])):
                        row_warnings.append("Btw-modus onbekend voor een of meer prijskolommen.")
                    if not has_identity:
                        row_warnings.append("Product gebruikt fallback-identiteit zonder artikelnummer/EAN/supplierCode.")
                    preview_rows.append(
                        preview_row(
                            source_path,
                            analysis_path,
                            sheet_name,
                            row_number,
                            "product",
                            headers,
                            values,
                            section_label,
                            normalized,
                            row_warnings,
                        )
                    )
                    rows.append(normalized)
                else:
                    preview_rows.append(
                        preview_row(
                            source_path,
                            analysis_path,
                            sheet_name,
                            row_number,
                            "ignored",
                            headers,
                            values,
                            section_label,
                            None,
                            ["Rij kon niet naar een productregel worden genormaliseerd."],
                        )
                    )
    finally:
        workbook.close()
    return rows, preview_rows, warnings


def summarize(
    rows: list[dict[str, Any]],
    preview_rows: list[dict[str, Any]],
    warnings: list[str],
    source_file_count: int,
) -> dict[str, Any]:
    by_file = Counter(row["sourceFileName"] for row in rows)
    by_category = Counter(row["categoryName"] for row in rows)
    price_count = sum(len(row.get("prices", [])) for row in rows)
    unknown_vat = sum(
        1
        for row in rows
        for price in row.get("prices", [])
        if price.get("vatMode") == "unknown"
    )
    row_kind_counts = Counter(row.get("rowKind", "unknown") for row in preview_rows)
    status_counts = Counter(row.get("status", "unknown") for row in preview_rows)
    zero_or_no_price_rows = sum(
        1 for row in preview_rows if zero_or_no_price_warning(row.get("warnings", []))
    )
    warning_rows = sum(1 for row in preview_rows if row.get("warnings"))
    error_rows = sum(1 for row in preview_rows if row.get("errors"))
    unknown_vat_rows = sum(
        1
        for row in preview_rows
        if any("btw-modus onbekend" in warning.lower() for warning in row.get("warnings", []))
    )

    return {
        "tenantSlug": "henke-wonen",
        "productRows": len(rows),
        "previewRows": len(preview_rows),
        "priceRules": price_count,
        "unknownVatModePriceRules": unknown_vat,
        "unknownVatModeRows": unknown_vat_rows,
        "sourceFiles": {
            "count": source_file_count,
            "withProductRows": len(by_file),
        },
        "categories": dict(by_category.most_common()),
        "sourceFileName": dict(by_file.most_common()),
        "rowKinds": dict(row_kind_counts.most_common()),
        "statuses": dict(status_counts.most_common()),
        "warnings": {
            "count": warning_rows,
            "errors": error_rows,
            "duplicateOrSkippedExactCopies": warnings,
            "zeroOrNoPriceRows": zero_or_no_price_rows,
        },
        "vatMapping": {
            "unresolvedVatMappings": None,
            "note": "Ontbrekende btw-mappings komen uit Convex importprofielen en worden in de portal bewaakt.",
        },
    }


def build_markdown(summary: dict[str, Any]) -> str:
    lines = [
        "# Catalogusimport samenvatting",
        "",
        "`docs/catalog-import-summary.md` is de primaire reviewbron voor de catalogusvoorvertoning.",
        "De grote rij-preview wordt standaard niet gegenereerd. Gebruik `npm run catalog:preview -- --full` of `CATALOG_PREVIEW_FULL=1` alleen voor debug/ontwikkeling.",
        "",
        "## Samenvatting",
        "",
        f"- Productregels: {summary['productRows']}",
        f"- Voorvertonings-/auditregels: {summary['previewRows']}",
        f"- Prijsregels: {summary['priceRules']}",
        f"- Prijsregels met onbekende btw-modus: {summary['unknownVatModePriceRules']}",
        f"- Rijen met waarschuwing over onbekende btw-modus: {summary['unknownVatModeRows']}",
        f"- Bronbestanden totaal: {summary['sourceFiles']['count']}",
        f"- Bronbestanden met productregels: {summary['sourceFiles']['withProductRows']}",
        f"- Rijen zonder bruikbare prijs: {summary['warnings']['zeroOrNoPriceRows']}",
        f"- Ontbrekende btw-mappings: {summary['vatMapping']['unresolvedVatMappings'] if summary['vatMapping']['unresolvedVatMappings'] is not None else 'niet beschikbaar in lokale Excel-voorvertoning'}",
        "",
        "## ZTAHL btw-bron",
        "",
        "- De btw-bevestiging voor ZTAHL komt uit de Excel print-header, niet uit een cel.",
        "- `Verkoopprijslijst ZTAHL 2026 - NL.xlsx`: `ZTAHL verkoopprijslijst incl. BTW - 2026`.",
        "- `D-Inkoopprijslijst ZTAHL 2026 - NL.xlsx`: `ZTAHL inkooppprijslijst excl. BTW - 2026`.",
        "",
        "## Per categorie",
        "",
        "| Categorie | Rijen |",
        "| --- | ---: |",
    ]
    for name, count in summary["categories"].items():
        lines.append(f"| {name} | {count} |")
    lines.extend(["", "## Per bronbestand", "", "| Bestand | Productregels |", "| --- | ---: |"])
    for name, count in summary["sourceFileName"].items():
        lines.append(f"| `{name}` | {count} |")
    lines.extend(["", "## Regeltypes", "", "| Regeltype | Rijen |", "| --- | ---: |"])
    for name, count in summary["rowKinds"].items():
        lines.append(f"| {label_row_kind(name)} | {count} |")
    lines.extend(["", "## Statussen", "", "| Status | Rijen |", "| --- | ---: |"])
    for name, count in summary["statuses"].items():
        lines.append(f"| {label_status(name)} | {count} |")
    if summary["warnings"]["duplicateOrSkippedExactCopies"]:
        lines.extend(["", "## Dubbele/overgeslagen exacte kopieën", ""])
        for warning in summary["warnings"]["duplicateOrSkippedExactCopies"]:
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Volledige preview",
            "",
            "- Standaard wordt geen rij-previewbestand geschreven.",
            f"- Debugpad: `{FULL_ROWS_OUT.relative_to(ROOT)}`",
            "- Volledige output is JSONL: één auditbare JSON-regel per voorvertoningsregel.",
        ]
    )
    lines.append("")
    return "\n".join(lines)


def build_sample(rows: list[dict[str, Any]], preview_rows: list[dict[str, Any]]) -> str:
    warning_error_rows = [
        row for row in preview_rows if row.get("warnings") or row.get("errors")
    ][:25]
    unknown_vat_rows = [
        row
        for row in preview_rows
        if any("btw-modus onbekend" in warning.lower() for warning in row.get("warnings", []))
    ][:25]
    lines = [
        "# Catalogusimport voorbeeld",
        "",
        "Compact voorbeeldbestand voor snelle menselijke controle. Maximaal 25 regels per sectie.",
        "",
        "## Eerste 25 productregels",
        "",
        "| Bestand | Tabblad | Rij | Product | Categorie | Prijsregels |",
        "| --- | --- | ---: | --- | --- | ---: |",
    ]
    for row in rows[:25]:
        product = str(row.get("productName", "")).replace("|", "\\|")
        lines.append(
            f"| `{row.get('sourceFileName')}` | `{row.get('sourceSheetName')}` | {row.get('sourceRowNumber')} | {product} | {row.get('categoryName')} | {len(row.get('prices', []))} |"
        )
    lines.extend(
        [
            "",
            "## Eerste 25 waarschuwings-/foutregels",
            "",
            "| Bestand | Tabblad | Rij | Regeltype | Status | Meldingen |",
            "| --- | --- | ---: | --- | --- | --- |",
        ]
    )
    for row in warning_error_rows:
        messages = "; ".join(row.get("warnings", []) + row.get("errors", [])).replace("|", "\\|")
        lines.append(
            f"| `{row.get('sourceFileName')}` | `{row.get('sourceSheetName')}` | {row.get('rowNumber')} | {label_row_kind(row.get('rowKind'))} | {label_status(row.get('status'))} | {messages} |"
        )
    lines.extend(
        [
            "",
            "## Eerste 25 regels met onbekende btw-modus",
            "",
            "| Bestand | Tabblad | Rij | Product | Meldingen |",
            "| --- | --- | ---: | --- | --- |",
        ]
    )
    for row in unknown_vat_rows:
        normalized = row.get("normalized") or {}
        product = str(normalized.get("productName", "")).replace("|", "\\|")
        messages = "; ".join(row.get("warnings", [])).replace("|", "\\|")
        lines.append(
            f"| `{row.get('sourceFileName')}` | `{row.get('sourceSheetName')}` | {row.get('rowNumber')} | {product} | {messages} |"
        )
    lines.append("")
    return "\n".join(lines)


def write_full_preview(rows: list[dict[str, Any]], preview_rows: list[dict[str, Any]]) -> None:
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    with FULL_ROWS_OUT.open("w", encoding="utf-8") as handle:
        for row in preview_rows:
            handle.write(
                json.dumps(
                    {"type": "previewRow", "tenantSlug": "henke-wonen", **row},
                    ensure_ascii=False,
                    sort_keys=True,
                )
                + "\n"
            )
        for row in rows:
            handle.write(
                json.dumps(
                    {"type": "productRow", "tenantSlug": "henke-wonen", **row},
                    ensure_ascii=False,
                    sort_keys=True,
                )
                + "\n"
            )


def cli_options(argv: list[str]) -> dict[str, Any]:
    options: dict[str, Any] = {
        "fullPreview": os.environ.get("CATALOG_PREVIEW_FULL") == "1",
        "noWrite": False,
        "sourceFilters": [],
    }
    index = 0
    while index < len(argv):
        arg = argv[index]
        if arg == "--full":
            options["fullPreview"] = True
        elif arg == "--no-write":
            options["noWrite"] = True
        elif arg == "--source":
            index += 1
            if index >= len(argv):
                raise SystemExit("--source requires a file-name or relative-path filter")
            options["sourceFilters"].append(argv[index].lower())
        elif arg.startswith("--source="):
            options["sourceFilters"].append(arg.split("=", 1)[1].lower())
        index += 1
    return options


def matches_source_filter(path: Path, filters: list[str]) -> bool:
    if not filters:
        return True
    rel = str(path.relative_to(ROOT)).replace("\\", "/").lower()
    name = path.name.lower()
    return any(filter_value in name or filter_value in rel for filter_value in filters)


def main() -> None:
    options = cli_options(sys.argv[1:])
    full_preview = options["fullPreview"]
    no_write = options["noWrite"]
    source_filters = options["sourceFilters"]
    if not no_write:
        OUT_DIR.mkdir(exist_ok=True)
    flexcolours_files = sorted(
        path
        for path in DATA_DIR.rglob("*")
        if path.is_file()
        and path.suffix.lower() in {".xlsx", ".xls"}
        and is_flexcolours(path)
    )
    if flexcolours_files and not source_filters:
        import sys as _sys
        print(
            f"[INFO] {len(flexcolours_files)} FlexColours raambekleding bestand(en) overgeslagen "
            "(prijsmatrix-structuur; vereist aparte parser). Bestanden:",
            file=_sys.stderr,
        )
        for _fc in flexcolours_files:
            print(f"  - {_fc.relative_to(ROOT)}", file=_sys.stderr)
    source_files = sorted(
        path
        for path in DATA_DIR.rglob("*")
        if path.is_file()
        and (
            (path.suffix.lower() in {".xlsx", ".xls"} and not is_flexcolours(path))
            or (path.suffix.lower() == ".pdf" and is_hebeta(path))
        )
        and path.name != "Henke Wonen Jeffrey.xlsx"
        and matches_source_filter(path, source_filters)
    )
    if not source_files:
        raise SystemExit("No source files matched the catalog preview filters.")
    rows: list[dict[str, Any]] = []
    preview_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    seen_hashes: dict[str, Path] = {}
    for source_path in source_files:
        file_hash = sha256(source_path)
        if file_hash in seen_hashes:
            warnings.append(
                f"{source_path.relative_to(ROOT)} is een exacte kopie van {seen_hashes[file_hash].relative_to(ROOT)} en is overgeslagen."
            )
            continue
        seen_hashes[file_hash] = source_path
        parsed_rows, parsed_preview_rows, parsed_warnings = parse_workbook(source_path)
        rows.extend(parsed_rows)
        preview_rows.extend(parsed_preview_rows)
        warnings.extend(parsed_warnings)

    summary = summarize(rows, preview_rows, warnings, len(source_files))
    if not no_write:
        SUMMARY_JSON_OUT.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        SUMMARY_OUT.write_text(build_markdown(summary), encoding="utf-8")
        SAMPLE_OUT.write_text(build_sample(rows, preview_rows), encoding="utf-8")
        # Write the full import payload — required by upload_catalog_batch_import.mjs.
        # Only written when no source filter is active so a partial run never leaves a
        # stale / incomplete preview.json on disk.
        if not source_filters:
            PREVIEW_JSON_OUT.write_text(
                json.dumps(
                    {
                        "tenantSlug": summary["tenantSlug"],
                        "rows": rows,
                        "previewRows": preview_rows,
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
        if full_preview:
            write_full_preview(rows, preview_rows)
    print(
        json.dumps(
            {
                "sourceFilter": source_filters or None,
                "writeOutputs": not no_write,
                "sourceFiles": summary["sourceFiles"]["count"],
                "rows": summary["productRows"],
                "previewRows": summary["previewRows"],
                "prices": summary["priceRules"],
                "unknownVatModePriceRules": summary["unknownVatModePriceRules"],
                "categories": summary["categories"],
                "warnings": warnings[:20],
                "summary": str(SUMMARY_OUT.relative_to(ROOT)) if not no_write else None,
                "summaryJson": str(SUMMARY_JSON_OUT.relative_to(ROOT)) if not no_write else None,
                "sample": str(SAMPLE_OUT.relative_to(ROOT)) if not no_write else None,
                "fullPreview": str(FULL_ROWS_OUT.relative_to(ROOT)) if full_preview and not no_write else None,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
