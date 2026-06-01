"""
FlexColours raambekleding matrix-parser v2.

Modelleert elke (bestand, sheet, prijsgroep) combinatie als ÉÉN product
met de volledige breedte×hoogte prijsmatrix opgeslagen in attributes.
Dit resulteert in ~45 producten i.p.v. 29.060.

Gebruik:
    node tools/run_python_tool.mjs tools/parse_flexcolours_matrix.py
    node tools/run_python_tool.mjs tools/parse_flexcolours_matrix.py --no-write
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
FLEX_DIR = ROOT / "DATA" / "Leveranciers" / "Unilin Flooring" / "FlexColours"
OUT_DIR = ROOT / "docs"
PREVIEW_JSON_OUT = OUT_DIR / "flexcolours-import-preview.json"
SUMMARY_OUT = OUT_DIR / "flexcolours-import-summary.json"

SUPPLIER_NAME = "Unilin Flooring"
BRAND_NAME = "FlexColours"
CATEGORY_SLUG = "raambekleding"
CATEGORY_NAME = "Raambekleding"
TENANT_SLUG = "henke-wonen"

PRIJSGROEP_RE = re.compile(r"prijsgroep\s*([A-Za-z0-9]+)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def stable_hash(parts: list[str]) -> str:
    text = "\x00".join(str(p) for p in parts)
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:32]


def cell_val(cell) -> str | None:
    if not hasattr(cell, "value") or cell.value is None:
        return None
    s = str(cell.value).strip()
    return s if s else None


def to_int(text: str | None) -> int | None:
    if not text:
        return None
    try:
        return int(float(text))
    except (ValueError, TypeError):
        return None


def to_float(text: str | None) -> float | None:
    if not text:
        return None
    try:
        return float(text)
    except (ValueError, TypeError):
        return None


def is_prijsgroep_label(text: str | None) -> bool:
    return bool(text and PRIJSGROEP_RE.search(text))


def extract_prijsgroep(text: str) -> str:
    m = PRIJSGROEP_RE.search(text)
    return m.group(1) if m else text.strip()


def is_numeric_str(text: str | None) -> bool:
    if not text:
        return False
    try:
        float(text)
        return True
    except (ValueError, TypeError):
        return False


def extract_lamelbreedte(sheet_name: str) -> str | None:
    m = re.search(r"(\d+)\s*mm", sheet_name, re.IGNORECASE)
    return f"{m.group(1)} mm" if m else None


def product_kind_for(file_name: str, sheet_name: str) -> str:
    fn, sn = file_name.lower(), sheet_name.lower()
    if "plisse" in sn or "plisse" in fn:
        return "plisse"
    if "duette" in sn or "duette" in fn:
        return "duette"
    if "verticale" in fn or "vertical" in fn:
        return "jaloezie"
    if "horizontaal" in fn or "horizontal" in fn:
        return "jaloezie"
    if "hout" in fn or "geweven" in fn:
        return "jaloezie"
    return "blind"


def product_type_label(file_name: str, sheet_name: str) -> str:
    """Mensleesbaar label voor het producttype (voor naam opbouw)."""
    fn = file_name.replace(" - 2026.xlsx", "").replace(".xlsx", "").strip()
    sn = sheet_name.strip()
    if sn.lower() not in fn.lower():
        return f"{fn} – {sn}"
    return fn


# ---------------------------------------------------------------------------
# Sheet parser — levert MatrixBlock objecten
# ---------------------------------------------------------------------------

class MatrixBlock:
    """Één prijsgroep-blok binnen een sheet: breedte-as, hoogte-as, prijsmatrix."""

    def __init__(self, prijsgroep: str, file_name: str, sheet_name: str):
        self.prijsgroep = prijsgroep
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.widths: list[int] = []         # kolom A waarden (breedtes)
        self.heights: list[int] = []        # kolomheader waarden (hoogtes)
        self._col_map: dict[int, int] = {}  # col_index → hoogte_index
        self._current_row: list[float | None] = []
        self._rows: list[list[float | None]] = []  # rows[breedte_idx][hoogte_idx]

    def set_height_headers(self, col_height_pairs: list[tuple[int, int]]) -> None:
        self.heights = [h for _, h in col_height_pairs]
        self._col_map = {col: idx for idx, (col, _) in enumerate(col_height_pairs)}

    def add_matrix_row(self, breedte: int, row_data: dict[int, str]) -> None:
        self.widths.append(breedte)
        price_row: list[float | None] = []
        for col_idx in sorted(self._col_map):
            p = to_float(row_data.get(col_idx))
            price_row.append(p if (p is not None and p > 0) else None)
        self._rows.append(price_row)

    def min_price(self) -> float | None:
        for row in self._rows:
            for p in row:
                if p is not None:
                    return p
        return None

    def is_valid(self) -> bool:
        return bool(self.widths and self.heights and self._rows)

    def to_attributes(self, lamelbreedte: str | None) -> dict[str, Any]:
        attrs: dict[str, Any] = {
            "prijsgroep": self.prijsgroep,
            "widths": self.widths,
            "heights": self.heights,
            "matrix": self._rows,
        }
        if lamelbreedte:
            attrs["lamelBreedte"] = lamelbreedte
        return attrs


def parse_sheet_blocks(
    source_path: Path,
    sheet_name: str,
    ws,
) -> list[MatrixBlock]:
    """
    Leest één sheet en retourneert een lijst van MatrixBlocks
    (één per PRIJSGROEP-sectie).
    """
    blocks: list[MatrixBlock] = []
    current: MatrixBlock | None = None
    expecting_heights = False
    max_row = ws.max_row or 0

    for r_num, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row), start=1):
        row_data: dict[int, str] = {}
        for cell in row:
            v = cell_val(cell)
            if v is not None and hasattr(cell, "column"):
                row_data[cell.column] = v

        if not row_data:
            continue

        col_a = row_data.get(1)

        # ── PRIJSGROEP-label ───────────────────────────────────────────────
        if col_a and is_prijsgroep_label(col_a):
            # Sla vorige block op als valide
            if current and current.is_valid():
                blocks.append(current)
            current = MatrixBlock(extract_prijsgroep(col_a), source_path.name, sheet_name)
            expecting_heights = True
            continue

        if current is None:
            continue

        # ── Hoogte-header-rij: col A leeg, rest zijn numerieke hoogtes ────
        is_height_row = (
            (expecting_heights and col_a is None)
            or col_a == "-"
        ) and any(is_numeric_str(row_data.get(c)) for c in row_data if c > 1)

        if is_height_row:
            pairs = []
            for col_idx in sorted(k for k in row_data if k > 1):
                h = to_int(row_data[col_idx])
                if h and h > 0:
                    pairs.append((col_idx, h))
            current.set_height_headers(pairs)
            expecting_heights = False
            continue

        expecting_heights = False

        # ── Matrix-rij: col A is breedte ──────────────────────────────────
        if not current.heights or not is_numeric_str(col_a):
            continue

        breedte = to_int(col_a)
        if breedte and breedte > 0:
            current.add_matrix_row(breedte, row_data)

    # Laatste block
    if current and current.is_valid():
        blocks.append(current)

    return blocks


# ---------------------------------------------------------------------------
# Product row builder
# ---------------------------------------------------------------------------

def block_to_product_row(block: MatrixBlock, source_path: Path) -> dict[str, Any]:
    lamelbreedte = extract_lamelbreedte(block.sheet_name)
    type_label = product_type_label(block.file_name, block.sheet_name)
    product_name = f"{type_label} – Prijsgroep {block.prijsgroep}"
    kind = product_kind_for(block.file_name, block.sheet_name)

    import_key = stable_hash([
        SUPPLIER_NAME,
        CATEGORY_SLUG,
        block.file_name,
        block.sheet_name,
        block.prijsgroep,
    ])

    attrs = block.to_attributes(lamelbreedte)

    # Minimumprijs als "vanaf"-prijs
    min_price = block.min_price()
    prices = []
    if min_price is not None:
        source_key = f"{import_key}:vanaf"
        prices.append({
            "sourceKey": source_key,
            "priceType": "retail",
            "priceUnit": "piece",
            "vatMode": "unknown",
            "currency": "EUR",
            "amount": min_price,
            "sourceColumnName": "Vanafprijs (laagste dimensie)",
        })

    return {
        "importKey": import_key,
        "sourceFileName": block.file_name,
        "sourceSheetName": block.sheet_name,
        "sourcePath": str(source_path.relative_to(ROOT)),
        "sourceRowNumber": 1,
        "supplierName": SUPPLIER_NAME,
        "brandName": BRAND_NAME,
        "categorySlug": CATEGORY_SLUG,
        "categoryName": CATEGORY_NAME,
        "productName": product_name,
        "productKind": kind,
        "productType": "made_to_measure",
        "sectionLabel": f"Prijsgroep {block.prijsgroep}",
        "unit": "piece",
        "widthMm": block.widths[0] if block.widths else None,   # min breedte
        "lengthMm": block.heights[0] if block.heights else None, # min hoogte
        "attributes": attrs,
        "prices": prices,
    }


# ---------------------------------------------------------------------------
# File parser
# ---------------------------------------------------------------------------

def parse_file(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            blocks = parse_sheet_blocks(path, sheet_name, ws)
            for block in blocks:
                rows.append(block_to_product_row(block, path))
    finally:
        wb.close()
    return rows


# ---------------------------------------------------------------------------
# Preview-rij builder
# ---------------------------------------------------------------------------

def to_preview_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "rowKind": "product",
        "status": "warning",
        "sourceFileName": row["sourceFileName"],
        "sourceSheetName": row.get("sourceSheetName"),
        "rowNumber": row["sourceRowNumber"],
        "rowHash": row["importKey"],
        "raw": row,
        "normalized": row,
        "warnings": ["Btw-modus onbekend. Matrix-prijzen staan in attributes.matrix."],
        "errors": [],
    }


# ---------------------------------------------------------------------------
# Summary & CLI
# ---------------------------------------------------------------------------

def build_summary(rows: list[dict[str, Any]], files: list[Path]) -> dict[str, Any]:
    from collections import Counter
    by_file: Counter[str] = Counter(r["sourceFileName"] for r in rows)
    by_pg: Counter[str] = Counter(r.get("sectionLabel", "?") for r in rows)
    return {
        "tenantSlug": TENANT_SLUG,
        "sourceFiles": len(files),
        "productRows": len(rows),
        "priceRules": sum(len(r.get("prices", [])) for r in rows),
        "note": "Elk product bevat een volledige breedte×hoogte matrix in attributes.",
        "byFile": dict(by_file.most_common()),
        "byPrijsgroep": dict(by_pg.most_common()),
    }


def cli_options(argv: list[str]) -> dict[str, Any]:
    opts: dict[str, Any] = {"noWrite": False}
    for arg in argv:
        if arg == "--no-write":
            opts["noWrite"] = True
    return opts


def main() -> None:
    opts = cli_options(sys.argv[1:])
    no_write = opts["noWrite"]

    source_files = sorted(FLEX_DIR.glob("*.xlsx"))
    if not source_files:
        raise SystemExit(f"Geen FlexColours bestanden gevonden in {FLEX_DIR}")

    if not no_write:
        OUT_DIR.mkdir(exist_ok=True)

    all_rows: list[dict[str, Any]] = []
    for path in source_files:
        file_rows = parse_file(path)
        all_rows.extend(file_rows)
        print(
            json.dumps({"file": path.name, "products": len(file_rows)}, ensure_ascii=False),
            file=sys.stderr,
        )

    preview_rows = [to_preview_row(r) for r in all_rows]
    summary = build_summary(all_rows, source_files)

    if not no_write:
        PREVIEW_JSON_OUT.write_text(
            json.dumps({"tenantSlug": TENANT_SLUG, "rows": all_rows, "previewRows": preview_rows}, ensure_ascii=False),
            encoding="utf-8",
        )
        SUMMARY_OUT.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    print(json.dumps({**summary, "writeOutputs": not no_write}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
