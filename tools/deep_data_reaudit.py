from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from audit_excel_data import (
    CODE_HEADER_RE,
    DATA_DIR,
    ROOT,
    clean_text,
    detect_header_row,
    infer_quantity_kind,
    is_price_header,
    is_quantity_header,
    row_kind,
    sheet_rows,
)
from build_catalog_import import (
    ATTRIBUTE_HEADERS,
    build_row,
    converted_xlsx_for,
    headers_for,
    key,
    normalize,
    sha256,
)

OUT_DIR = ROOT / "docs"
OUT_JSON = OUT_DIR / "data-reaudit-2026-04-29.json"
OUT_MD = OUT_DIR / "data-reaudit-2026-04-29.md"
PREVIEW_JSON = OUT_DIR / "catalog-import-preview.json"

PRODUCT_NAME_HEADERS = {
    "kwaliteitsnaam",
    "kwaliteit",
    "omschrijving",
    "material description",
    "decor name",
    "afmeting (cm)",
    "quality",
    "design",
    "type",
}

IDENTITY_HEADERS = {
    "artikelnummer",
    "art.nr.",
    "supplier code",
    "sap codes floors",
    "commercial code",
    "ean",
    "ean code",
    "ean-code",
    "sku",
}

DIMENSION_HEADERS = {
    "afmeting",
    "afmetingen",
    "breedte (cm)",
    "breedte",
    "breedte in cm",
    "width",
    "w",
    "lengte (cm)",
    "lengte in cm",
    "lengte plank (cm)",
    "lengte mm",
    "lengte",
    "length",
    "l",
    "dikte (mm)",
    "totale dikte (mm)",
    "totaal dikte (mm)",
    "dikte mm",
    "dikte in mm",
    "dikte",
    "toplaag",
    "toplaag mm",
    "toplaag (mm)",
    "tick",
    "dikte toplaag (mm)",
    "toplaag in mm",
    "aantal m2 per pak",
    "pakinhoud",
    "pakinhoud (m²)",
    "m2",
    "aantal panelen per pak",
    "planken per pak",
    "panels",
    "aantal pakken per pallet",
    "aantal pakker per pallet",
    "pakken per pallet",
    "packs",
}

COMMERCIAL_ALIAS_HEADERS = {
    "ambiant collectie",
    "ambiant kleur",
    "floorlife collectie",
    "floorlife kleur",
}

OTHER_MAPPED_HEADERS = {
    "kleur",
    "kleurindicatie",
    "kleurnummer",
    "artikelgroep",
    "soort",
    "company",
    "collection",
    "aantal kleuren",
}

SUSPICIOUS_UNMAPPED_PATTERNS = re.compile(
    r"(btw|vat|prijs|price|code|ean|art|collect|kleur|color|breedte|width|lengte|length|"
    r"dikte|pak|pack|pallet|trailer|afname|bestel|unit|eenheid|materiaal|material|"
    r"kamerhoog|lining|pattern|weight|washing|composition|roman|panel|curtain|mart|"
    r"garantie|project|rug|backing|warmte|vocht)",
    re.IGNORECASE,
)


def canonical(header: str) -> str:
    return normalize(header).strip()


def classify_header(header: str) -> str:
    normalized = canonical(header)
    if not normalized:
        return "empty"
    if is_price_header(header):
        return "price"
    if CODE_HEADER_RE.search(header) or normalized in IDENTITY_HEADERS:
        return "identity"
    if normalized in PRODUCT_NAME_HEADERS:
        return "product_name"
    if normalized in DIMENSION_HEADERS:
        return "dimension_or_package"
    if normalized in COMMERCIAL_ALIAS_HEADERS:
        return "commercial_alias"
    if normalized in OTHER_MAPPED_HEADERS:
        return "mapped_metadata"
    if is_quantity_header(header):
        return "quantity_or_logistics"
    if normalized in ATTRIBUTE_HEADERS or any(token in normalized for token in ATTRIBUTE_HEADERS):
        return "attribute"
    return "unmapped"


def source_files() -> list[Path]:
    return sorted(
        path
        for path in DATA_DIR.rglob("*")
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xls"}
    )


def analyze_header_coverage() -> dict[str, Any]:
    header_counts: Counter[str] = Counter()
    header_class_counts: Counter[str] = Counter()
    unmapped_counts: Counter[str] = Counter()
    suspicious_unmapped: Counter[str] = Counter()
    quantity_headers: dict[str, str] = {}
    row_compare: list[dict[str, Any]] = []
    import_rows_without_prices: list[dict[str, Any]] = []
    seen_hashes: dict[str, Path] = {}

    for source_path in source_files():
        analysis_path = converted_xlsx_for(source_path)
        if analysis_path is None:
            continue
        file_hash = sha256(source_path)
        if file_hash in seen_hashes:
            row_compare.append(
                {
                    "sourcePath": str(source_path.relative_to(ROOT)),
                    "sheetName": "*",
                    "auditProductRows": 0,
                    "normalizedRows": 0,
                    "delta": 0,
                    "note": f"Exacte kopie van {seen_hashes[file_hash].relative_to(ROOT)}; overgeslagen in importpreview.",
                }
            )
            continue
        seen_hashes[file_hash] = source_path
        workbook = load_workbook(analysis_path, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                header_row, headers = headers_for(ws)
                if not header_row or not headers:
                    row_compare.append(
                        {
                            "sourcePath": str(source_path.relative_to(ROOT)),
                            "sheetName": sheet_name,
                            "auditProductRows": 0,
                            "normalizedRows": 0,
                            "delta": 0,
                            "note": "Geen header gevonden.",
                        }
                    )
                    continue
                max_col = len(headers)
                code_indexes = [
                    index for index, header in enumerate(headers) if CODE_HEADER_RE.search(header)
                ]
                price_indexes = [
                    index for index, header in enumerate(headers) if is_price_header(header)
                ]
                audit_product_rows = 0
                normalized_rows = 0
                section_label = None
                empty_streak = 0
                start = header_row + 1
                max_row = min(ws.max_row or 0, 50000)
                for row_number, row_values in enumerate(
                    sheet_rows(ws, start, max_row, max_col),
                    start=start,
                ):
                    values = list(row_values)
                    kind = row_kind(values, code_indexes, price_indexes)
                    if kind == "empty":
                        empty_streak += 1
                        if row_number > start + 100 and empty_streak >= 500:
                            break
                        continue
                    empty_streak = 0
                    if kind == "section":
                        section_label = clean_text(next((value for value in values if clean_text(value)), ""))
                        continue
                    if kind != "product":
                        continue
                    audit_product_rows += 1
                    normalized = build_row(
                        source_path,
                        analysis_path,
                        sheet_name,
                        headers,
                        values,
                        row_number,
                        section_label,
                        file_hash,
                    )
                    if normalized:
                        has_identity = any(
                            normalized.get(field)
                            for field in ["articleNumber", "supplierCode", "commercialCode", "ean"]
                        )
                        if not normalized.get("prices") and not has_identity:
                            section_label = normalized["productName"]
                            continue
                        normalized_rows += 1
                        if not normalized.get("prices"):
                            import_rows_without_prices.append(
                                {
                                    "sourcePath": str(source_path.relative_to(ROOT)),
                                    "sheetName": sheet_name,
                                    "rowNumber": row_number,
                                    "productName": normalized.get("productName"),
                                    "categoryName": normalized.get("categoryName"),
                                }
                            )
                    for index, header in enumerate(headers):
                        if index >= len(values) or not header:
                            continue
                        if clean_text(values[index]) == "":
                            continue
                        header_counts[header] += 1
                        header_class = classify_header(header)
                        header_class_counts[header_class] += 1
                        if header_class == "unmapped":
                            unmapped_counts[header] += 1
                            if SUSPICIOUS_UNMAPPED_PATTERNS.search(header):
                                suspicious_unmapped[header] += 1
                        if header_class == "quantity_or_logistics":
                            quantity_headers[header] = infer_quantity_kind(header)
                row_compare.append(
                    {
                        "sourcePath": str(source_path.relative_to(ROOT)),
                        "sheetName": sheet_name,
                        "auditProductRows": audit_product_rows,
                        "normalizedRows": normalized_rows,
                        "delta": audit_product_rows - normalized_rows,
                    }
                )
        finally:
            workbook.close()

    return {
        "headerClassCounts": dict(header_class_counts),
        "topUnmappedHeaders": unmapped_counts.most_common(80),
        "topSuspiciousUnmappedHeaders": suspicious_unmapped.most_common(80),
        "quantityHeaders": quantity_headers,
        "rowCompare": row_compare,
        "importRowsWithoutPrices": import_rows_without_prices[:300],
    }


def analyze_preview() -> dict[str, Any]:
    payload = json.loads(PREVIEW_JSON.read_text(encoding="utf-8"))
    rows = payload.get("rows", [])
    price_type_counts = Counter()
    price_unit_counts = Counter()
    vat_mode_counts = Counter()
    category_counts = Counter()
    supplier_counts = Counter()
    source_counts = Counter()
    source_sheet_counts = Counter()
    unique_import_keys = set()
    price_list_groups: dict[tuple[str, str], set[tuple[str, str]]] = defaultdict(set)
    duplicate_identity_groups: Counter[str] = Counter()
    future_validity_rows = []
    for row in rows:
        category_counts[row.get("categoryName", "Onbekend")] += 1
        supplier_counts[row.get("supplierName", "Onbekend")] += 1
        source_counts[row.get("sourcePath", row.get("sourceFileName", "Onbekend"))] += 1
        source_sheet_counts[
            f"{row.get('sourcePath', row.get('sourceFileName', 'Onbekend'))}::{row.get('sourceSheetName', '')}"
        ] += 1
        unique_import_keys.add(row.get("importKey"))
        duplicate_identity_groups[
            "|".join(
                [
                    row.get("supplierName", ""),
                    row.get("categoryName", ""),
                    row.get("articleNumber")
                    or row.get("supplierCode")
                    or row.get("commercialCode")
                    or row.get("ean")
                    or row.get("productName", ""),
                ]
            )
        ] += 1
        validity_text = json.dumps(
            {
                "sourceFileName": row.get("sourceFileName"),
                "prices": row.get("prices", []),
            },
            ensure_ascii=False,
        ).lower()
        if "vanaf" in validity_text or re.search(r"\d{1,2}[./-]\d{1,2}[./-]20\d{2}", validity_text):
            future_validity_rows.append(
                {
                    "sourceFileName": row.get("sourceFileName"),
                    "sourceSheetName": row.get("sourceSheetName"),
                    "productName": row.get("productName"),
                    "sourceRowNumber": row.get("sourceRowNumber"),
                    "hasValidity": bool(row.get("year") or row.get("validFrom")),
                    "pricesMissingValidity": sum(
                        1
                        for price in row.get("prices", [])
                        if (
                            "vanaf" in str(price.get("sourceColumnName", "")).lower()
                            or re.search(r"\d{1,2}[./-]\d{1,2}[./-]20\d{2}", str(price.get("sourceColumnName", "")))
                        )
                        and not price.get("validFrom")
                    ),
                }
            )
        for price in row.get("prices", []):
            price_type_counts[price.get("priceType", "unknown")] += 1
            price_unit_counts[price.get("priceUnit", "unknown")] += 1
            vat_mode_counts[price.get("vatMode", "unknown")] += 1
        price_list_groups[(row.get("sourceFileName", ""), row.get("sourceSheetName", ""))].add(
            (row.get("sourcePath", ""), row.get("fileHash", ""))
        )

    merged_price_list_risks = []
    for (file_name, sheet_name), sources in price_list_groups.items():
        source_paths = {source for source, _hash in sources}
        hashes = {file_hash for _source, file_hash in sources}
        if len(source_paths) > 1:
            merged_price_list_risks.append(
                {
                    "sourceFileName": file_name,
                    "sourceSheetName": sheet_name,
                    "sourcePathCount": len(source_paths),
                    "fileHashCount": len(hashes),
                    "sourcePaths": sorted(source_paths),
                }
            )

    duplicate_identities = [
        {"identity": identity, "rows": count}
        for identity, count in duplicate_identity_groups.most_common()
        if count > 1
    ][:80]

    return {
        "rows": len(rows),
        "prices": sum(len(row.get("prices", [])) for row in rows),
        "uniqueImportKeys": len({key for key in unique_import_keys if key}),
        "categoryCounts": dict(category_counts.most_common()),
        "supplierCounts": dict(supplier_counts.most_common()),
        "sourceCounts": dict(source_counts.most_common()),
        "priceTypeCounts": dict(price_type_counts.most_common()),
        "priceUnitCounts": dict(price_unit_counts.most_common()),
        "vatModeCounts": dict(vat_mode_counts.most_common()),
        "mergedPriceListRisks": merged_price_list_risks,
        "duplicateIdentities": duplicate_identities,
        "futureValiditySamples": future_validity_rows[:40],
    }


def build_markdown(report: dict[str, Any]) -> str:
    preview = report["preview"]
    coverage = report["coverage"]
    lines = [
        "# Henke Wonen Data Reaudit - 2026-04-29",
        "",
        "Doel: de ruwe Exceldata opnieuw naast de huidige Convex/importcode leggen en expliciet markeren waar nog data verloren kan gaan.",
        "",
        "## Kerncijfers",
        "",
        f"- Excelbestanden in `DATA`: {report['fileCount']}",
        f"- Sheets: {report['sheetCount']}",
        f"- Genormaliseerde catalogusrijen: {preview['rows']}",
        f"- Genormaliseerde prijsregels: {preview['prices']}",
        f"- Unieke importKeys in preview: {preview['uniqueImportKeys']}",
        f"- Prijsregels met `vatMode=unknown`: {preview['vatModeCounts'].get('unknown', 0)}",
        f"- Genormaliseerde rijen zonder prijs: {len(coverage['importRowsWithoutPrices'])}",
        "",
        "## Belangrijkste nieuwe/blijvende aandachtspunten",
        "",
    ]
    findings = report["findings"]
    for finding in findings:
        lines.append(f"- **{finding['severity']}** - {finding['title']}: {finding['detail']}")

    lines.extend(["", "## Rijdekking per sheet", "", "| Bron | Sheet | Audit productrijen | Genormaliseerd | Delta |", "| --- | --- | ---: | ---: | ---: |"])
    for item in coverage["rowCompare"]:
        if item["auditProductRows"] or item["normalizedRows"] or item.get("note"):
            lines.append(
                f"| `{item['sourcePath']}` | `{item['sheetName']}` | {item['auditProductRows']} | "
                f"{item['normalizedRows']} | {item['delta']} |"
            )

    lines.extend(["", "## Prijstypes", "", "| Type | Aantal |", "| --- | ---: |"])
    for name, count in preview["priceTypeCounts"].items():
        lines.append(f"| `{name}` | {count} |")

    lines.extend(["", "## Prijseenheden", "", "| Eenheid | Aantal |", "| --- | ---: |"])
    for name, count in preview["priceUnitCounts"].items():
        lines.append(f"| `{name}` | {count} |")

    lines.extend(["", "## Btw-modus", "", "| vatMode | Aantal |", "| --- | ---: |"])
    for name, count in preview["vatModeCounts"].items():
        lines.append(f"| `{name}` | {count} |")

    lines.extend(["", "## Top ongemapte headers", "", "| Header | Gevulde cellen |", "| --- | ---: |"])
    for header, count in coverage["topUnmappedHeaders"][:40]:
        lines.append(f"| `{header}` | {count} |")

    lines.extend(["", "## Verdachte ongemapte headers", "", "| Header | Gevulde cellen |", "| --- | ---: |"])
    for header, count in coverage["topSuspiciousUnmappedHeaders"][:40]:
        lines.append(f"| `{header}` | {count} |")

    lines.extend(["", "## PriceList samenvoeg-risico", ""])
    if preview["mergedPriceListRisks"]:
        for item in preview["mergedPriceListRisks"]:
            lines.append(
                f"- `{item['sourceFileName']}` / `{item['sourceSheetName']}` komt uit "
                f"{item['sourcePathCount']} paden en {item['fileHashCount']} hashes."
            )
            for source in item["sourcePaths"]:
                lines.append(f"  - `{source}`")
    else:
        lines.append("- Geen bronpad-duplicaten met dezelfde bestandsnaam/sheet gevonden.")

    lines.extend(["", "## Rijen zonder prijs in normalisatie", ""])
    if coverage["importRowsWithoutPrices"]:
        for item in coverage["importRowsWithoutPrices"][:50]:
            lines.append(
                f"- `{item['sourcePath']}` / `{item['sheetName']}` rij {item['rowNumber']}: "
                f"{item['productName']} ({item['categoryName']})"
            )
    else:
        lines.append("- Geen genormaliseerde catalogusrijen zonder prijs.")

    lines.extend(["", "## Categorieen in preview", "", "| Categorie | Rijen |", "| --- | ---: |"])
    for name, count in preview["categoryCounts"].items():
        lines.append(f"| {name} | {count} |")

    lines.append("")
    return "\n".join(lines)


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    files = source_files()
    coverage = analyze_header_coverage()
    preview = analyze_preview()
    sheet_count = len(coverage["rowCompare"])
    findings = []

    if preview["mergedPriceListRisks"]:
        findings.append(
            {
                "severity": "HOOG",
                "title": "priceLists worden nog te grof herkend",
                "detail": "catalogImport.ensurePriceList zoekt op sourceFileName + sourceSheetName. Duplicaten in andere paden kunnen daardoor dezelfde priceList delen; fileHash/sourcePath moet in de sleutel.",
            }
        )
    if preview["vatModeCounts"].get("unknown", 0):
        findings.append(
            {
                "severity": "HOOG",
                "title": "btw-modus is meestal onbekend",
                "detail": f"{preview['vatModeCounts'].get('unknown', 0)} prijsregels hebben vatMode=unknown. Definitieve import moet mapping blokkeren of expliciet laten bevestigen.",
            }
        )
    if coverage["importRowsWithoutPrices"]:
        findings.append(
            {
                "severity": "MIDDEL",
                "title": "catalogusrijen zonder prijs",
                "detail": f"{len(coverage['importRowsWithoutPrices'])} rijen normaliseren wel als product maar krijgen geen prijsregel; deze moeten in preview als warning zichtbaar zijn.",
            }
        )
    quantity_header_names = set(coverage["quantityHeaders"].keys())
    if "Aantal kleuren" in quantity_header_names:
        findings.append(
            {
                "severity": "MIDDEL",
                "title": "`Aantal kleuren` wordt als logistieke quantity gezien",
                "detail": "Dit is waarschijnlijk collectie-metadata, geen order/logistiek veld. Nu verdwijnt dit uit attributes doordat quantityheaders worden uitgesloten.",
            }
        )
    suspicious_unmapped = [
        item for item in coverage["topSuspiciousUnmappedHeaders"] if item[0].lower() != "artikel"
    ]
    if suspicious_unmapped:
        headers = ", ".join(f"`{header}`" for header, _ in suspicious_unmapped[:8])
        findings.append(
            {
                "severity": "MIDDEL",
                "title": "verdachte headers vallen buiten mapping",
                "detail": f"Vooral {headers}. Controleer of dit echte eigenschappen of importprofielvelden moeten zijn.",
            }
        )
    if any(
        not item.get("hasValidity") or item.get("pricesMissingValidity", 0)
        for item in preview["futureValiditySamples"]
    ):
        findings.append(
            {
                "severity": "MIDDEL",
                "title": "geldigheidsdatums uit bestandsnamen/headers worden nog niet gezet",
                "detail": "Roots vanaf 01-05-2026 en 2026-collecties worden wel geïmporteerd, maar validFrom/year worden niet structureel afgeleid in de cataloguspreview.",
            }
        )
    if preview["rows"] != preview["uniqueImportKeys"]:
        findings.append(
            {
                "severity": "INFO",
                "title": "deduplicatie verklaart lager productaantal",
                "detail": f"{preview['rows']} previewrijen leveren {preview['uniqueImportKeys']} unieke importKeys op; Convex productaantal lager dan rijaantal is dus verwacht, maar prijzen moeten per priceList/sourceKey blijven bestaan.",
            }
        )

    report = {
        "fileCount": len(files),
        "sheetCount": sheet_count,
        "preview": preview,
        "coverage": coverage,
        "findings": findings,
    }
    OUT_JSON.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    OUT_MD.write_text(build_markdown(report), encoding="utf-8")
    print(
        json.dumps(
            {
                "fileCount": len(files),
                "sheetCount": sheet_count,
                "rows": preview["rows"],
                "prices": preview["prices"],
                "findings": findings,
            },
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
